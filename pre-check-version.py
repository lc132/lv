#!/usr/bin/env python3
# 读取策略调整记录（获取 file_version 和 params）
# 依赖外部函数：safe_read_json, read_all_history, log_alert, os, beijing_date
import os

BUILTIN_VERSION = "v6.13.36"  # 与 ashare_screener.py 保持一致
adj_records = safe_read_json('/workspace/策略调整记录.json')
if adj_records and len(adj_records) > 0:
    latest = adj_records[-1]
    file_version = latest.get('version', BUILTIN_VERSION)
    params = latest.get('params', {})
else:
    file_version = BUILTIN_VERSION
    params = {}

# 读取所有日期归档的推荐历史找最后一个strategy_check
history = read_all_history()
last_check = None
current_version = None  # 显式初始化，防止首次运行 NameError
for r in reversed(history):
    if r.get('type') == 'strategy_check':
        last_check = r
        break
if last_check:
    current_version = last_check.get('version', 'unknown')
    if current_version != file_version:
        log_alert("INFO", "版本检查", f"推荐历史版本{current_version}≠策略调整版本{file_version}，以策略调整为准")
    else:
        log_alert("INFO", "版本检查", f"版本一致{file_version}")

# 自动更新筛选条件表格
if last_check is None or current_version != file_version:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        xlsx_path = "/workspace/A股短线选股筛选条件.xlsx"
        if os.path.exists(xlsx_path):
            wb = load_workbook(xlsx_path)
            _cell_font = Font(name='Arial', size=10)
            _bold_font = Font(name='Arial', size=10, bold=True)
            _thin_border = Border(
                left=Side(style='thin', color='B0B0B0'),
                right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'),
                bottom=Side(style='thin', color='B0B0B0'),
            )

            def _wc(ws, r, c, v, font=_cell_font):
                # 注意：此函数与步骤26(十三.A)中的 _wc 定义一致，修改时需同步更新两处
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                        if not (r == mr.min_row and c == mr.min_col):
                            return
                        ws.unmerge_cells(str(mr))
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = font
                cell.border = _thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            # 更新筛选条件概述
            ws1 = wb['筛选条件概述']
            _wc(ws1, 1, 1, f'A股短线选股筛选条件 — {file_version}', _bold_font)
            _wc(ws1, 2, 2, file_version)
            _wc(ws1, 2, 3, f'{beijing_date}更新')
            # 追加版本记录
            vr = ws1.max_row + 1
            _wc(ws1, vr, 1, file_version)
            _wc(ws1, vr, 2, beijing_date)
            _wc(ws1, vr, 3, '版本同步')

            # 更新关键纪律版本号
            if '关键纪律' in wb.sheetnames:
                ws11 = wb['关键纪律']
                _wc(ws11, 1, 1, f'关键纪律 — {file_version}', _bold_font)

            wb.save(xlsx_path)
            wb.close()
            log_alert("INFO", "筛选条件", f"筛选条件.xlsx 已同步至 {file_version}")
        else:
            log_alert("WARNING", "筛选条件", "筛选条件.xlsx 不存在，跳过自动更新")
    except Exception as e:
        log_alert("WARNING", "筛选条件", f"筛选条件.xlsx 自动更新失败: {str(e)[:80]}")
