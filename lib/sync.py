#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
步骤23-28: 回溯检查昨日做T、告警日志摘要、筛选概况、GitHub同步、飞书推送、每周复盘拉取
"""
from lib.core import *

# ============================================================
# 步骤23: 回溯检查昨日做T
# ============================================================
def step23_backtrack_do_T(ctx):
    print("\n" + "=" * 60)
    print("步骤23: 回溯检查昨日做T")
    print("=" * 60)
    
    all_history = ctx.get('all_history', [])
    data_date = ctx['data_date']
    yesterday = (datetime.strptime(data_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # 查找昨日的 do_T_eval
    yesterday_evals = []
    for r in all_history:
        if r.get('type') == 'do_T_eval' and r.get('date') == yesterday:
            yesterday_evals.append(r)
    
    if not yesterday_evals:
        print(f"  昨日({yesterday})无做T评估记录")
        return
    
    # 查找昨日的 do_T 执行记录
    yesterday_do_T = []
    for r in all_history:
        if r.get('type') == 'do_T' and r.get('date') == yesterday:
            yesterday_do_T.append(r.get('code', ''))
    
    reminded = 0
    for ev in yesterday_evals:
        feasible = ev.get('do_T_feasible', False)
        code = ev.get('code', '?')
        name = ev.get('name', '?')
        
        if feasible in (True, '谨慎', 'True') and code not in yesterday_do_T:
            msg = f"  ⚠️ {code} {name}: 昨日做T评估为{feasible}，但未执行做T"
            print(msg)
            reminded += 1
        elif feasible == '观望':
            print(f"  {code} {name}: 昨日评估为观望，无需操作")
        elif feasible == False:
            print(f"  {code} {name}: 昨日评估为不可做T，无需操作")
        elif code in yesterday_do_T:
            print(f"  {code} {name}: 昨日已执行做T ✅")
    
    if reminded == 0:
        print(f"  无需提醒")
    
    ctx['_do_T_backtrack_reminded'] = reminded

# ============================================================
# 步骤24: 告警日志摘要
# ============================================================
def step24_alert_summary(ctx):
    print("\n" + "=" * 60)
    print("步骤24: 告警日志摘要")
    print("=" * 60)
    
    alert_path = f"{DATA_DIR}/系统告警.log"
    try:
        with open(alert_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        today_lines = [l for l in lines if ctx['data_date'] in l]
        if today_lines:
            print(f"  今日告警: {len(today_lines)}条")
            for l in today_lines[-5:]:
                print(f"    {l.strip()}")
        else:
            print("  今日无异常")
    except FileNotFoundError:
        print("  今日无异常")

# ============================================================
# 步骤25: 输出筛选概况
# ============================================================
def step25_summary(ctx):
    print("\n" + "=" * 60)
    print("📊 筛选概况")
    print("=" * 60)
    
    strategy_counts = ctx.get('strategy_counts', Counter())
    
    summary = f"""
📊 筛选概况 — {ctx['prediction_date']} (数据来源: {ctx['data_date']})
① 原始标的池: {ctx.get('total_raw', 0)}只 → ② 硬排除: {ctx.get('excluded_count', 0)}只 → ③ 信号过滤: {ctx.get('signal_dropped', 0)}只 → ④ 策略匹配: {ctx.get('passed_strategy', 0)}只 → ⑤ 行业限制: {ctx.get('passed_industry', 0)}只 → ⑥ 新闻筛查: {ctx.get('passed_news', 0)}只 → ★ 最终: {ctx.get('final_recommend_count', 0)}只
策略分布: A:{strategy_counts.get('A',0)} B:{strategy_counts.get('B',0)} C:{strategy_counts.get('C',0)} D:{strategy_counts.get('D',0)} E:{strategy_counts.get('E',0)}
市场环境: {ctx.get('market_condition', '未知')} | 建议仓位: {ctx.get('position', 0)}%
"""
    
    if ctx.get('exclusion_stats'):
        print("排除TOP5:")
        for reason, count in ctx['exclusion_stats'].most_common(5):
            print(f"  {reason}: {count}只")
    
    print(summary)
    ctx['summary'] = summary

# ============================================================
# 步骤26: GitHub同步
# ============================================================
def step26_github_sync(ctx):
    print("\n" + "=" * 60)
    print("步骤26: GitHub同步")
    print("=" * 60)
    
    md_path = ctx.get('md_path', '')
    html_path = ctx.get('html_path', '')
    report_dir = ctx.get('report_dir', '')
    prediction_date = ctx['prediction_date']
    pred_yyyymmdd = prediction_date.replace('-', '')
    
    if not os.path.exists(md_path):
        log_alert("WARNING", "GitHub同步", "md文件不存在，跳过")
        print("  md文件不存在，跳过")
        return
    
    repo_url = f"https://github.com/{GITHUB_REPO}.git"
    repo_dir = f"{TEMP_DIR}/lv_sync"
    
    try:
        if os.path.exists(repo_dir):
            shutil.rmtree(repo_dir, ignore_errors=True)
        
        # 使用 _git_with_token 安全传递 Token（避免 Token 出现在进程列表中）
        # 注意：此函数定义在 ashare_screener.py 中，lib 模块通过 ctx 调用
        _git_with_token = ctx.get('_git_with_token')
        if _git_with_token:
            _git_with_token(
                ["git", "clone", "--depth", "1", "--branch", "main", repo_url, repo_dir],
                timeout=30
            )
        else:
            subprocess.run(
                ["git", "clone", "--depth", "1", "--branch", "main", repo_url, repo_dir],
                capture_output=True, text=True, timeout=30, check=True
            )
        
        # 校验筛选条件xlsx版本（SKILL §十三.A: 推送前检查版本一致性）
        local_xlsx = f"{DATA_DIR}/A股短线选股筛选条件.xlsx"
        try:
            from openpyxl import load_workbook
            wb = load_workbook(local_xlsx, read_only=True)
            # 尝试从第一Sheet A1读取版本号
            xlsx_version = None
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                a1_val = ws.cell(row=1, column=1).value
                if a1_val and 'v' in str(a1_val).lower():
                    xlsx_version = str(a1_val).strip()
            wb.close()
            file_version = ctx.get('file_version', '')
            if xlsx_version and file_version and xlsx_version != file_version:
                log_alert("INFO", "GitHub同步", f"筛选条件xlsx版本({xlsx_version})≠代码版本({file_version})，同步更新")
                shutil.copy(local_xlsx, os.path.join(repo_dir, "A股短线选股筛选条件.xlsx"))
                print(f"  筛选条件xlsx版本不一致({xlsx_version}≠{file_version})，已同步更新")
        except FileNotFoundError:
            log_alert("INFO", "GitHub同步", "筛选条件xlsx不存在，跳过版本检查")
        except Exception:
            log_alert("INFO", "GitHub同步", "筛选条件xlsx版本检查跳过（文件不可读）")
        
        # 清理超15天旧文件
        cutoff_date = datetime.now() - timedelta(days=15)
        for f in os.listdir(repo_dir):
            if f.startswith("推荐历史_") and f.endswith(".json"):
                try:
                    date_str = f.replace("推荐历史_", "").replace(".json", "")
                    f_date = datetime.strptime(date_str, '%Y%m%d')
                    if f_date < cutoff_date:
                        os.remove(os.path.join(repo_dir, f))
                        print(f"  清理旧文件: {f}")
                except Exception:
                    pass
            if (f.startswith("短线标的_") and f.endswith(".md")):
                try:
                    date_str = f.replace("短线标的_", "").replace(".md", "").replace("-", "")
                    f_date = datetime.strptime(date_str, '%Y%m%d')
                    if f_date < cutoff_date:
                        os.remove(os.path.join(repo_dir, f))
                        print(f"  清理旧文件: {f}")
                except Exception:
                    pass
        
        for f in os.listdir(repo_dir):
            if f.startswith("ashare-screening-"):
                try:
                    date_str = f.replace("ashare-screening-", "")
                    f_date = datetime.strptime(date_str, '%Y%m%d')
                    if f_date < cutoff_date:
                        shutil.rmtree(os.path.join(repo_dir, f), ignore_errors=True)
                        print(f"  清理旧目录: {f}")
                except Exception:
                    pass
        
        # 复制文件
        shutil.copy(md_path, os.path.join(repo_dir, f"短线标的_{prediction_date}.md"))
        
        # 推送持仓跟踪
        local_holding = f"{DATA_DIR}/持仓跟踪.xlsx"
        if os.path.exists(local_holding):
            shutil.copy(local_holding, os.path.join(repo_dir, "持仓跟踪.xlsx"))
        
        # 推送推荐历史归档
        local_archive = f"{DATA_DIR}/推荐历史_{pred_yyyymmdd}.json"
        if os.path.exists(local_archive):
            shutil.copy(local_archive, os.path.join(repo_dir, f"推荐历史_{pred_yyyymmdd}.json"))
        
        # 推送HTML报告
        if os.path.exists(html_path):
            dest_html_dir = os.path.join(repo_dir, f"ashare-screening-{pred_yyyymmdd}")
            if os.path.exists(dest_html_dir):
                shutil.rmtree(dest_html_dir, ignore_errors=True)
            shutil.copytree(report_dir, dest_html_dir)
        
        # 推送 GitHub Pages 首页 index.html
        local_index = f"{DATA_DIR}/index.html"
        if os.path.exists(local_index):
            shutil.copy(local_index, os.path.join(repo_dir, "index.html"))
        
        # 推送 SKILL.md 规则文件（保持仓库版本与当前运行版本一致）
        local_skill = f"{DATA_DIR}/SKILL.md"
        if not os.path.exists(local_skill):
            local_skill = "/workspace/SKILL.md"
        if os.path.exists(local_skill):
            shutil.copy(local_skill, os.path.join(repo_dir, "SKILL.md"))
            print(f"  已同步 SKILL.md 规则文件")
        
        # 版本变更时同步推送策略调整记录
        if ctx.get('_version_changed'):
            local_adjust = f"{DATA_DIR}/策略调整记录.json"
            if os.path.exists(local_adjust):
                shutil.copy(local_adjust, os.path.join(repo_dir, "策略调整记录.json"))
                print(f"  版本变更 → 同步策略调整记录")
                log_alert("INFO", "GitHub同步", f"版本变更，同步策略调整记录")
        
        # Git操作
        subprocess.run(["git", "-C", repo_dir, "config", "user.email", "ashare-bot@github.com"], check=True, timeout=10)
        subprocess.run(["git", "-C", repo_dir, "config", "user.name", "ashare-screener"], check=True, timeout=10)
        subprocess.run(["git", "-C", repo_dir, "add", "-A"], check=True, timeout=10)
        
        commit_msg = f"筛选结果 {prediction_date}"
        result = subprocess.run(
            ["git", "-C", repo_dir, "commit", "-m", commit_msg],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            if "nothing to commit" in (result.stderr + result.stdout):
                log_alert("INFO", "GitHub同步", "无变更，跳过推送")
                return
            log_alert("WARNING", "GitHub同步", f"commit失败: {result.stderr[:100]}")
            return
        
        # 如果有变更则推送（使用 _git_with_token 安全传递 Token）
        _git_with_token = ctx.get('_git_with_token')
        if _git_with_token:
            push_result = _git_with_token(
                ["git", "-C", repo_dir, "push", "origin", "main"],
                timeout=60, check=False
            )
        else:
            push_result = subprocess.run(
                ["git", "-C", repo_dir, "push", "origin", "main"],
                capture_output=True, text=True, timeout=30
            )
        
        if push_result.returncode == 0:
            print(f"  ✅ GitHub同步成功: {prediction_date}")
            log_alert("INFO", "GitHub同步", f"✅ {prediction_date} 已推送")
        else:
            print(f"  ⚠️ 推送: {push_result.stderr[:100]}")
            log_alert("WARNING", "GitHub同步", f"推送结果: {push_result.stderr[:100]}")
        
    except Exception as e:
        log_alert("WARNING", "GitHub同步", f"失败: {str(e)[:100]}")
        print(f"  ⚠️ GitHub同步失败: {str(e)[:80]}")
    finally:
        if os.path.exists(repo_dir):
            shutil.rmtree(repo_dir, ignore_errors=True)

# ============================================================
# 步骤27: 飞书推送（单卡片，GitHub Pages 跳转链接）
# ============================================================
def step27_feishu_push(ctx):
    print("\n" + "=" * 60)
    print("步骤27: 飞书推送")
    print("=" * 60)
    
    if not FEISHU_WEBHOOK:
        log_alert("WARNING", "飞书推送", "未配置Webhook URL，跳过")
        print("  未配置Webhook，跳过")
        return
    
    strategy_counts = ctx.get('strategy_counts', Counter())
    prediction_date = ctx['prediction_date']
    pred_yyyymmdd = prediction_date.replace('-', '')
    html_path = ctx.get('html_path', '')
    
    # GitHub Pages URL（仓库已启用 Pages，从 main 分支根目录服务）
    pages_base = f"https://lc132.github.io/lv"
    pages_home = pages_base
    pages_report = f"{pages_base}/ashare-screening-{pred_yyyymmdd}/ashare-screening-{pred_yyyymmdd}.html"
    
    final_count = ctx.get('final_recommend_count', 0)
    
    # 构建推荐标的摘要（仅策略分布，详细信息在 Pages 报告中）
    strategy_lines = []
    for s in ['A', 'B', 'C', 'D', 'E']:
        cnt = strategy_counts.get(s, 0)
        if cnt > 0:
            names = {'A': '动量延续', 'B': '超跌反弹', 'C': '事件驱动', 'D': '回调企稳', 'E': '资金埋伏'}
            strategy_lines.append(f"{names[s]}({s}): {cnt}只")
    strategy_summary = "  ".join(strategy_lines) if strategy_lines else "无"
    
    card = {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"📊 每日短线标的筛选 — {prediction_date}"},
                "template": "blue"
            },
            "elements": [
                {
                    "tag": "div",
                    "text": {
                        "tag": "lark_md",
                        "content": f"**数据来源**: {ctx['data_date']}  |  **市场环境**: {ctx.get('market_condition','未知')}  |  **建议仓位**: {ctx.get('position',0)}%"
                    }
                },
                {"tag": "hr"},
                {
                    "tag": "div",
                    "text": {
                        "tag": "lark_md",
                        "content": f"原始标的池: **{ctx.get('total_raw',0)}**只 → 硬排除: **{ctx.get('excluded_count',0)}**只 → 信号过滤: **{ctx.get('signal_dropped',0)}**只 → 策略匹配: **{ctx.get('passed_strategy',0)}**只 → 行业限制: **{ctx.get('passed_industry',0)}**只 → 新闻筛查: **{ctx.get('passed_news',0)}**只 → ★ 最终: **{final_count}**只"
                    }
                },
                {"tag": "hr"},
                {
                    "tag": "div",
                    "text": {
                        "tag": "lark_md",
                        "content": f"**策略分布**: {strategy_summary}"
                    }
                },
                {"tag": "hr"},
                {
                    "tag": "div",
                    "text": {
                        "tag": "lark_md",
                        "content": f"📈 [**查看完整可视化报告（GitHub Pages）**]({pages_report})"
                    }
                },
                {
                    "tag": "div",
                    "text": {
                        "tag": "lark_md",
                        "content": f"📁 [**报告列表首页**]({pages_home})"
                    }
                },
                {"tag": "note", "elements": [{"tag": "plain_text", "content": "⚠️ 仅供参考，不构成投资建议"}]}
            ]
        }
    }
    
    try:
        req = urllib.request.Request(
            FEISHU_WEBHOOK,
            data=json.dumps(card, ensure_ascii=False).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read())
        if result.get('code') == 0:
            print(f"  ✅ 飞书推送成功（GitHub Pages: {pages_report}）")
            log_alert("INFO", "飞书推送", f"✅ {prediction_date} 已推送到飞书群（Pages: {pages_report}）")
        else:
            print(f"  ⚠️ 飞书推送失败: {result.get('msg','')}")
            log_alert("WARNING", "飞书推送", f"推送失败: {result.get('msg','')}")
    except Exception as e:
        print(f"  ⚠️ 飞书推送异常: {str(e)[:80]}")
        log_alert("WARNING", "飞书推送", f"请求异常: {str(e)[:100]}")

# ============================================================
# 步骤28: 每周复盘拉取（仅周六执行）
# ============================================================
def step28_weekly_review(ctx):
    """每周六从GitHub拉取本周推荐文件，汇总复盘"""
    print("\n" + "=" * 60)
    print("步骤28: 每周复盘拉取")
    print("=" * 60)
    
    weekday = ctx.get('beijing_weekday', 0)
    if weekday != 5:  # 0=Mon ... 5=Sat
        print(f"  非周六（周{weekday+1}），跳过复盘")
        return
    
    print("  周六 → 执行每周复盘...")
    
    if not GITHUB_TOKEN: print("  ⚠️ 无GitHub Token，跳过复盘"); return  # v6.12.8
    repo_url = f"https://{GITHUB_TOKEN}@github.com/{GITHUB_REPO}.git"
    temp_dir = f"{TEMP_DIR}/lv_weekly_review"
    
    try:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        
        subprocess.run(
            ["git", "clone", "--depth", "1", "--branch", "main", repo_url, temp_dir],
            capture_output=True, text=True, timeout=60, check=True
        )
        
        # 收集本周推荐的MD文件
        md_files = []
        for f in os.listdir(temp_dir):
            if f.startswith("短线标的_") and f.endswith(".md"):
                md_files.append(f)
        
        if not md_files:
            print("  本周无推荐文件，跳过复盘")
            log_alert("INFO", "每周复盘", "本周无推荐文件，跳过")
            return
        
        md_files.sort()
        print(f"  拉取到 {len(md_files)} 个推荐文件: {', '.join(md_files[-7:])}")
        
        # 汇总统计：推荐胜率、平均涨跌、策略分布
        total_recos = 0
        win_count = 0
        strategy_dist = Counter()
        all_changes = []
        
        for md_file in md_files:
            filepath = os.path.join(temp_dir, md_file)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                # 解析MD表格中的标的和涨幅
                lines = content.split('\n')
                in_table = False
                for line in lines:
                    stripped = line.strip()
                    if stripped.startswith('| 代码 ') or stripped.startswith('| 序号 '):
                        in_table = True
                        continue
                    if in_table and stripped.startswith('| '):
                        parts = [p.strip() for p in stripped.split('|') if p.strip()]
                        if len(parts) >= 6:
                            total_recos += 1
                            # 尝试解析策略和涨幅
                            try:
                                strategy_cell = parts[1] if len(parts) > 1 else ''
                                chg_cell = parts[5] if len(parts) > 5 else ''
                                # 策略列：检查中文名称
                                strategy_name_map = {'动量延续': 'A', '超跌反弹': 'B', '事件驱动': 'C', '回调企稳': 'D', '资金埋伏': 'E'}
                                for cn_name, s_key in strategy_name_map.items():
                                    if cn_name in strategy_cell:
                                        strategy_dist[s_key] += 1
                                        break
                                chg_str = chg_cell.replace('%', '').replace('+', '')
                                try:
                                    chg_val = float(chg_str)
                                    all_changes.append(chg_val)
                                    if chg_val > 2:
                                        win_count += 1
                                except ValueError:
                                    pass
                            except Exception:
                                pass
                    elif in_table and not stripped.startswith('| '):
                        in_table = False
            except Exception:
                continue
        
        avg_chg = round(sum(all_changes) / len(all_changes), 2) if all_changes else 0
        win_rate = round(win_count / total_recos * 100, 1) if total_recos > 0 else 0
        
        weekly_review = {
            "type": "weekly_review",
            "date": ctx['beijing_date'],
            "week_files": len(md_files),
            "total_recos": total_recos,
            "win_rate": win_rate,
            "avg_change": avg_chg,
            "strategy_dist": dict(strategy_dist),
        }
        safe_append_json(f"{DATA_DIR}/推荐历史_{ctx['beijing_date'].replace('-', '')}.json", weekly_review)
        
        print(f"  本周汇总: {len(md_files)}个推荐日, {total_recos}只标的")
        print(f"  兑现率(>2%): {win_rate}% | 平均涨跌: {avg_chg:+.2f}%")
        print(f"  策略分布: {dict(strategy_dist)}")
        log_alert("INFO", "每周复盘", f"本周{len(md_files)}日{total_recos}只, 胜率{win_rate}%, 均涨{avg_chg:+.2f}%")
        
    except Exception as e:
        log_alert("WARNING", "每周复盘", f"拉取失败: {str(e)[:100]}")
        print(f"  复盘拉取失败: {str(e)[:80]}")
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)