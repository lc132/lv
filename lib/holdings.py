#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
步骤4-4C: 持仓行情同步、做T评估、持仓跟踪同步、持仓危机检查
"""
from lib.core import *

# ============================================================
# 步骤4: 持仓行情同步
# ============================================================
def step4_holdings_sync(ctx):
    print("\n" + "=" * 60)
    print("步骤4: 持仓行情同步")
    print("=" * 60)
    
    holdings = ctx.get('holdings', [])
    if not holdings:
        print("  无持仓记录，跳过")
        return
    
    all_history = ctx.get('all_history', [])
    updated = 0
    
    for h in holdings:
        code = h.get('code', '')
        if not code:
            continue
        
        try:
            # 获取当日行情
            market = 'sz' if code.startswith(('000','002','003','300','301')) else 'sh'
            sina_url = f'https://hq.sinajs.cn/list={market}{code}'
            req = urllib.request.Request(sina_url, headers={
                'User-Agent': 'Mozilla/5.0',
                'Referer': 'https://finance.sina.com.cn'
            })
            resp = urllib.request.urlopen(req, timeout=5)
            text = resp.read().decode('gbk')
            
            if text and '=""' not in text:
                parts = text.split('"')[1].split(',')
                if len(parts) > 4:
                    prev_close = float(parts[2]) if parts[2] else 0
                    current = float(parts[3]) if parts[3] else 0
                    if current > 0:
                        # 保存旧值
                        old_current = h.get('current')
                        h['prev_close'] = prev_close
                        h['current'] = current
                        h['update_date'] = ctx['data_date']
                        # 计算盈亏
                        cost = h.get('cost', 0)
                        shares = h.get('shares', 0)
                        if cost > 0 and shares > 0:
                            h['pnl_pct'] = round((current - cost) / cost * 100, 2)
                            h['pnl_amount'] = round((current - cost) * shares, 2)
                            h['market_value'] = round(current * shares, 2)
                        updated += 1
                        print(f"  {code} {h.get('name','?')}: {old_current}→{current} (涨跌{h.get('pnl_pct',0)}%)")
        except Exception as e:
            log_alert("WARNING", "持仓行情同步", f"{code} 搜索失败: {str(e)[:60]}")
            continue
    
    # 更新回推荐历史文件（按日期分文件，找到含该holding的文件更新）
    if updated > 0:
        # 将更新后的holding记录写回对应日期的推荐历史文件
        data_yyyymmdd = ctx['data_date'].replace('-', '')
        hist_path = f"{DATA_DIR}/推荐历史_{data_yyyymmdd}.json"
        existing = safe_read_json(hist_path) or []
        
        for r in existing:
            if r.get('type') == 'holding':
                for h in holdings:
                    if h.get('code') == r.get('code'):
                        for k in ['current', 'prev_close', 'pnl_pct', 'pnl_amount', 'market_value', 'update_date']:
                            if k in h:
                                r[k] = h[k]
        
        safe_write_json(hist_path, existing)
        # 同时更新all_history
        for r in all_history:
            if r.get('type') == 'holding':
                for h in holdings:
                    if h.get('code') == r.get('code'):
                        for k in ['current', 'prev_close', 'pnl_pct', 'pnl_amount', 'market_value', 'update_date']:
                            if k in h:
                                r[k] = h[k]
        print(f"  已更新 {updated} 只持仓价格")
    
    ctx['holdings'] = holdings
    ctx['all_history'] = all_history

# ============================================================
# 步骤4A: 做T评估
# ============================================================
def step4A_do_T_eval(ctx):
    print("\n" + "=" * 60)
    print("步骤4A: 做T评估")
    print("=" * 60)
    
    holdings = ctx.get('holdings', [])
    if not holdings:
        print("  无持仓，跳过做T评估")
        return
    
    do_t_evals = []
    for h in holdings:
        pnl_pct = h.get('pnl_pct', 0)
        code = h.get('code', '?')
        name = h.get('name', '?')
        
        # Evaluate do T feasibility
        # 简化：按浮亏比例判断 + 连续成功/失败跟踪
        # 实际：需要下影/十字星/站MA5检测（K线历史数据目前不可达）
        do_t_feasible = '观望'
        pnl_ratio = abs(pnl_pct)
        
        if pnl_ratio < 1:
            do_t_feasible = False  # 浮亏<1%，无需
        elif pnl_ratio < 3:
            # 检查波动率是否≥3%（做T空间）
            amplitude = h.get('amplitude', 0)
            if amplitude >= 3:
                do_t_feasible = True
            else:
                do_t_feasible = '观望'  # 波动不足
        elif 3 <= pnl_ratio < 6:
            do_t_feasible = True
            # 检查前一日是否放量跌（volume_ratio>1.2→谨慎）
            vol_ratio = h.get('volume_ratio', 0)
            if vol_ratio and vol_ratio > 1.5:
                do_t_feasible = '谨慎'  # 放量跌→可能继续下行
                log_alert("INFO", "做T评估", f"{code} {name} 放量跌(量比{vol_ratio})，做T评估降为谨慎")
        else:
            do_t_feasible = '观望'  # 浮亏>6%，风险过大
        
        # 连续成功/失败计数跟踪
        do_t_history = [r for r in all_history if r.get('type') == 'do_T' and r.get('code') == code]
        recent_do_t = [r for r in do_t_history 
                       if datetime.strptime(r.get('date', '2000-01-01'), '%Y-%m-%d') >= 
                          (datetime.strptime(data_date, '%Y-%m-%d') - timedelta(days=30))]
        success_count = sum(1 for r in recent_do_t if r.get('result') == 'success')
        fail_count = sum(1 for r in recent_do_t if r.get('result') == 'fail')
        if fail_count >= 3 and success_count == 0:
            do_t_feasible = False
            log_alert("WARNING", "做T评估", f"{code} {name} 近30天做T连续{fail_count}次失败，暂停建议")
        
        print(f"  {code} {name}: {do_t_feasible} (浮亏{pnl_pct:.1f}%)")
        
        do_t_eval = {
            "type": "do_T_eval",
            "code": code,
            "name": name,
            "date": ctx['data_date'],
            "pnl_pct": pnl_pct,
            "do_T_feasible": do_t_feasible,
            "reason": f"浮亏{pnl_pct:.1f}%",
            "position_limit": 0
        }
        do_t_evals.append(do_t_eval)
    
    # 追加到推荐历史
    for eval_rec in do_t_evals:
        safe_append_json(f"{DATA_DIR}/推荐历史_{ctx['beijing_date'].replace('-', '')}.json", eval_rec)
    
    ctx['do_t_evals'] = do_t_evals

# ============================================================
# 步骤4B: 持仓跟踪同步到 xlsx
# ============================================================
def step4B_sync_holding_xlsx(ctx):
    """将步骤4更新后的持仓收盘价写入持仓跟踪.xlsx"""
    print("\n" + "=" * 60)
    print("步骤4B: 持仓跟踪同步")
    print("=" * 60)
    
    holdings = ctx.get('holdings', [])
    if not holdings:
        print("  无持仓记录，跳过")
        return
    
    xlsx_path = f"{DATA_DIR}/持仓跟踪.xlsx"
    if not os.path.exists(xlsx_path):
        log_alert("WARNING", "持仓跟踪同步", "xlsx不存在，跳过")
        print("  持仓跟踪.xlsx 不存在，跳过")
        return
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb["持仓明细"]
        
        # code → row mapping (skip header)
        code_row = {}
        for row in range(2, ws.max_row + 1):
            raw_code = ws.cell(row=row, column=1).value
            if raw_code:
                code = str(raw_code).strip()
                if len(code) == 4:
                    code = code.zfill(6)
                if code.isdigit() and len(code) == 6:
                    code_row[code] = row
        
        beijing_date = ctx['beijing_date']
        updated = 0
        for h in holdings:
            try:
                raw_code = h.get("code")
                code = str(raw_code) if raw_code is not None else ""
                current = h.get("current")
                if not code or code not in code_row:
                    if code:
                        log_alert("WARNING", "持仓跟踪同步", f"{code} 在xlsx中找不到")
                    continue
                if current is None:
                    log_alert("WARNING", "持仓跟踪同步", f"{code} 缺少current字段，跳过")
                    continue
                
                row = code_row[code]
                mv = h.get("market_value")
                pnl_amt = h.get("pnl_amount")
                if mv is None or pnl_amt is None:
                    cost = ws.cell(row=row, column=3).value
                    shares = ws.cell(row=row, column=4).value
                    if cost and shares and current:
                        mv = round(current * shares, 2)
                        pnl_amt = round((current - cost) * shares, 2)
                
                ws.cell(row=row, column=8).value = current  # H: 当前价
                if mv is not None:
                    ws.cell(row=row, column=9).value = mv  # I: 市值
                if pnl_amt is not None:
                    ws.cell(row=row, column=10).value = round(pnl_amt, 2)  # J: 盈亏额
                pnl_pct_val = h.get("pnl_pct")
                try:
                    pnl_pct_float = float(pnl_pct_val) if pnl_pct_val is not None else 0.0
                except (ValueError, TypeError):
                    pnl_pct_float = 0.0
                ws.cell(row=row, column=11).value = round(pnl_pct_float, 4)  # K: 盈亏率
                ws.cell(row=row, column=12).value = beijing_date  # L: 更新日期
                updated += 1
            except Exception as e:
                log_alert("WARNING", "持仓跟踪同步", f"单条异常(code={h.get('code','?')}): {str(e)[:80]}")
                continue
        
        if updated > 0:
            wb.save(xlsx_path)
            print(f"  已同步 {updated} 只持仓到持仓跟踪.xlsx")
            log_alert("INFO", "持仓跟踪同步", f"已更新{updated}只持仓价格")
        else:
            print(f"  无需更新（0只匹配）")
    except Exception as e:
        log_alert("WARNING", "持仓跟踪同步", f"失败: {str(e)[:100]}")
        print(f"  持仓跟踪同步失败: {str(e)[:80]}")

# ============================================================
# 步骤4C: 持仓危机检查
# ============================================================
def step4C_holding_crisis(ctx):
    print("\n" + "=" * 60)
    print("步骤4C: 持仓危机检查")
    print("=" * 60)
    
    holdings = ctx.get('holdings', [])
    alerts = []
    
    for h in holdings:
        code = h.get('code', '?')
        name = h.get('name', '?')
        cost = h.get('cost', 0)
        current = h.get('current', 0)
        prev_close = h.get('prev_close')
        pnl_pct = h.get('pnl_pct', 0)
        
        # 跌停检查
        if prev_close is not None and current > 0 and prev_close > 0:
            daily_chg = (current - prev_close) / prev_close * 100
            if daily_chg < -9.5:
                msg = f"⚠️ {code} {name} 当日跌停({daily_chg:.1f}%)！成本{cost} 现价{current} 浮亏{pnl_pct}%"
                alerts.append(msg)
                log_alert("WARNING", "持仓危机", msg)
        
        # 浮亏>15%
        if pnl_pct is not None and pnl_pct < -15:
            msg = f"⚠️ {code} {name} 浮亏突破15%做T上限({pnl_pct:.1f}%)，建议人工决策"
            alerts.append(msg)
            log_alert("WARNING", "持仓危机", msg)
        
        # L1硬排除检查
        if current > 0:
            l1_triggers = []
            if current < 5: l1_triggers.append("股价<5元(规则3)")
            if current > 100: l1_triggers.append("股价>100元(规则4)")
            if code.startswith("688"): l1_triggers.append("科创板(规则1)")
            if code.startswith("8") and len(str(code)) == 6: l1_triggers.append("北交所(规则2)")
            if l1_triggers:
                msg = f"⚠️ {code} {name} 触发L1硬排除: {', '.join(l1_triggers)}"
                alerts.append(msg)
                log_alert("WARNING", "持仓危机", msg)
    
    if alerts:
        for a in alerts:
            print(f"  {a}")
    else:
        print("  持仓无异常")
    
    ctx['holding_crisis_alerts'] = alerts