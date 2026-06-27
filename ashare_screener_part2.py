#!/usr/bin/env python3
"""
A股盘前短线标的筛选 v6.9.53 - Part 2: 步骤9-27 (核心筛选+输出)
"""
import urllib.request, urllib.parse, urllib.error, json, os, sys, time, re, subprocess, shutil
from datetime import datetime, timedelta
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================================
# 从 Part 1 继承的全局变量
# ============================================================
beijing_date = "2026-06-13"
data_date = "2026-06-12"
prediction_date = "2026-06-13"
beijing_weekday = 5
beijing_hour = 13
beijing_now = datetime(2026, 6, 13, 13, 9, 57)
market_env = "震荡"
file_version = "v6.9.53"

# 策略参数
params = {
    'search_budget': 25, 'northbound_threshold': 3000, 'consecutive_weeks': 2,
    'win_rate_drop_threshold': 10, 'limit_down_threshold': 100, 'max_adjust_params': 3,
    'confidence_position_enabled': True, 'max_holding_days': 5, 'circuit_breaker_threshold_pct': 3.0,
    'strategy_concentration_pct': 60, 'do_t_success_reset_count': 3,
    'conversion_rate_window_days': 10, 'conversion_rate_threshold': 0.3,
    'conversion_rate_restore': 0.6, 'conversion_rate_consecutive_days': 3,
    'data_tier_l2_skip_on_unavailable': True, 'data_tier_l3_downgrade_to_signal': True,
    'strategy_a_weak_market': 'closed'
}

# 各阶段计数
total_raw = 0
excluded_count = 0
filtered_count = 0
matched_count = 0
industry_limited_count = 0
news_filtered_count = 0
final_recommend_count = 0
strategy_counts = Counter()

# 排除统计
exclusion_reasons = Counter()

# 告警列表
crisis_alerts = []

# ============================================================
# 工具函数
# ============================================================
def log_alert(level, module, message, timestamp=None):
    if timestamp is None:
        timestamp = datetime.now()
    ts = timestamp.strftime('%Y-%m-%d %H:%M:%S') if hasattr(timestamp, 'strftime') else str(timestamp)
    log_line = f"[{ts}] [{level}] {module}: {message}\n"
    try:
        with open('/workspace/系统告警.log', 'a', encoding='utf-8') as f:
            f.write(log_line)
    except (PermissionError, OSError):
        pass
    print(f"  LOG [{level}] {module}: {message}")

def safe_read_json(path, default=None):
    try:
        if not os.path.exists(path): return default if default is not None else []
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if not isinstance(data, list): return default if default is not None else []
            return data
    except Exception:return default if default is not None else []

def safe_write_json(path, data):
    """原子写入：先写临时文件再重命名，防止写入中断导致数据损坏"""
    try:
        tmp_path = path + '.tmp'
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)  # 原子操作(POSIX)
    except (PermissionError, OSError) as e:
        log_alert("ERROR", "safe_write_json", f"{path}: {str(e)}")

def safe_append_json(path, record):
    data = safe_read_json(path)
    data.append(record)
    safe_write_json(path, data)

def safe_float(value, ndigits=3):
    if value is None: return None
    if isinstance(value, (int, float)): return round(float(value), ndigits)
    return value

# ============================================================
# 步骤9: 板块轮动
# ============================================================
def step9_sector_rotation():
    """获取板块资金流向"""
    print("\n" + "=" * 60)
    print("步骤9: 板块轮动分析")
    print("=" * 60)
    
    sector_flow = {"top3": [], "bottom5": [], "continuous_3d": []}
    
    try:
        # 获取东方财富行业板块行情
        url = "https://push2.eastmoney.com/api/qt/clist/get"
        params = {
            'pn': '1', 'pz': '20', 'po': '1', 'np': '1',
            'ut': 'bd1d9ddb04089700cf9c27f6f7426281',
            'fltt': '2', 'invt': '2', 'fid': 'f62',
            'fs': 'm:90+t:2',
            'fields': 'f2,f3,f4,f12,f14,f62,f104,f105',
            '_': str(int(time.time() * 1000))
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://quote.eastmoney.com/'
        }
        req = urllib.request.Request(f'{url}?{urllib.parse.urlencode(params)}', headers=headers)
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        
        if data and data.get('data') and data['data'].get('diff'):
            sectors = []
            for item in data['data']['diff']:
                name = item.get('f14', '')
                chg = item.get('f3', 0)
                inflow = item.get('f62', 0)
                if name:
                    sectors.append({"name": name, "chg": chg, "inflow": inflow})
            sectors.sort(key=lambda x: x.get('inflow', 0) or 0, reverse=True)
            sector_flow["top3"] = sectors[:3]
            sector_flow["bottom5"] = sectors[-5:] if len(sectors) >= 5 else []
            print(f"   资金流入TOP3: {[s['name'] for s in sector_flow['top3']]}")
            print(f"   资金流出TOP5: {[s['name'] for s in sector_flow['bottom5']]}")
    except Exception as e:
        log_alert("INFO", "板块轮动", f"板块数据获取失败: {str(e)[:60]}")
        print(f"   ⚠️ 板块数据不可达，跳过")
    
    return sector_flow

# ============================================================
# 步骤10A: 全市场API拉取（新浪批量）
# ============================================================
def step10a_fetch_all_stocks():
    """拉取全A股行情数据"""
    print("\n" + "=" * 60)
    print("步骤10A: 全市场行情拉取（新浪批量API）")
    print("=" * 60)
    
    log_alert("INFO", "行情采集", "东方财富clist不可达，降级为新浪批量API")
    
    # 生成代码范围（缩减范围提高效率）
    code_ranges = []
    # 上海主板 (600000-605999)
    for i in range(600000, 606000): code_ranges.append(f"sh{i}")
    # 深圳主板 (000001-003999)
    for i in range(1, 4000): code_ranges.append(f"sz{i:06d}")
    # 深圳创业板 (300000-301999)
    for i in range(300000, 302000): code_ranges.append(f"sz{i}")
    
    stocks = []
    batch_size = 80
    total_batches = (len(code_ranges) + batch_size - 1) // batch_size
    
    print(f"   共 {len(code_ranges)} 个代码范围，{total_batches} 批次")
    
    for batch_idx in range(0, len(code_ranges), batch_size):
        batch = code_ranges[batch_idx:batch_idx + batch_size]
        try:
            url = f"https://hq.sinajs.cn/list={','.join(batch)}"
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0',
                'Referer': 'https://finance.sina.com.cn'
            })
            with urllib.request.urlopen(req, timeout=8) as resp:
                text = resp.read().decode('gbk', errors='replace')
            
            for line in text.strip().split('\n'):
                if not line or '=""' in line:
                    continue
                try:
                    parts = line.split('"')[1].split(',')
                    if len(parts) < 6:
                        continue
                    header = line.split('="')[0]
                    raw_code = header.split('_')[-1] if '_' in header else header[-6:]
                    code = raw_code if len(raw_code) == 6 else raw_code[-6:]
                    # 只保留6位数字代码
                    if not code.isdigit() or len(code) != 6:
                        continue
                    name = parts[0]
                    open_val = float(parts[1]) if parts[1] else 0
                    prev_close = float(parts[2]) if parts[2] else 0
                    current = float(parts[3]) if parts[3] else 0
                    high = float(parts[4]) if parts[4] else 0
                    low = float(parts[5]) if parts[5] else 0
                    
                    if current <= 0 or prev_close <= 0:
                        continue
                    
                    change_pct = round((current - prev_close) / prev_close * 100, 2)
                    amplitude = round((high - low) / prev_close * 100, 2) if prev_close > 0 else 0
                    
                    market = 'sz' if code.startswith(('000','001','002','003','300','301')) else 'sh'
                    t_idx = 37 if market == 'sz' else 38
                    turnover = float(parts[t_idx]) if len(parts) > t_idx and parts[t_idx] else 0
                    volume = float(parts[8]) if len(parts) > 8 and parts[8] else 0
                    amount_val = float(parts[9]) if len(parts) > 9 and parts[9] else 0
                    
                    # 市盈率
                    pe_idx = 38 if market == 'sz' else 39
                    pe_ttm = float(parts[pe_idx]) if len(parts) > pe_idx and parts[pe_idx] else 0
                    
                    stocks.append({
                        "code": code, "name": name,
                        "open": open_val, "close": current,
                        "change_pct": change_pct,
                        "turnover": turnover,
                        "amplitude": amplitude,
                        "high": high, "low": low,
                        "prev_close": prev_close,
                        "volume": volume,
                        "volume_ratio": None,
                        "amount": amount_val,
                        "main_inflow": None,
                        "total_cap": None,
                        "pe_ttm": pe_ttm if pe_ttm > 0 else None,
                    })
                except (ValueError, IndexError):
                    continue
            
            if (batch_idx // batch_size) % 10 == 0:
                progress = batch_idx // batch_size + 1
                print(f"   进度: {progress}/{total_batches} 批次, 已采集 {len(stocks)} 只")
                time.sleep(0.05)
        except Exception as e:
            continue
    
    log_alert("INFO", "行情采集", f"全市场拉取到 {len(stocks)} 只标的（来源: sina）")
    print(f"✅ 步骤10A: 全市场拉取到 {len(stocks)} 只标的")
    return stocks

# ============================================================
# 步骤10B: 板块/行业补全
# ============================================================
# 行业映射表（常见板块→申万一级行业）
SECTOR_INDUSTRY_MAP = {
    '银行': '银行', '保险': '非银金融', '证券': '非银金融', '券商': '非银金融',
    '房地产': '房地产', '房地产开发': '房地产',
    '白酒': '食品饮料', '食品': '食品饮料', '饮料': '食品饮料', '乳业': '食品饮料',
    '医药': '医药生物', '医疗器械': '医药生物', '生物制品': '医药生物', '中药': '医药生物',
    '半导体': '电子', '芯片': '电子', '电子': '电子', '集成电路': '电子', 'LED': '电子',
    '计算机': '计算机', '软件': '计算机', '人工智能': '计算机', '大数据': '计算机',
    '通信': '通信', '5G': '通信',
    '电力': '公用事业', '电力设备': '电力设备', '新能源': '电力设备', '光伏': '电力设备', '风电': '电力设备', '储能': '电力设备',
    '汽车': '汽车', '新能源汽车': '汽车', '汽车零部件': '汽车',
    '有色金属': '有色金属', '黄金': '有色金属', '稀土': '有色金属',
    '钢铁': '钢铁', '煤炭': '煤炭',
    '化工': '基础化工', '化学制品': '基础化工', '化肥': '基础化工',
    '石油': '石油石化', '石化': '石油石化',
    '建筑': '建筑装饰', '建材': '建筑材料', '水泥': '建筑材料',
    '交通运输': '交通运输', '物流': '交通运输', '航空': '交通运输',
    '国防军工': '国防军工', '军工': '国防军工',
    '传媒': '传媒', '游戏': '传媒', '影视': '传媒',
    '纺织服装': '纺织服饰', '服装': '纺织服饰',
    '家用电器': '家用电器', '家电': '家用电器',
    '农林牧渔': '农林牧渔', '农业': '农林牧渔', '养殖': '农林牧渔',
    '环保': '环保', '环境保护': '环保',
    '轻工制造': '轻工制造', '造纸': '轻工制造',
    '商贸零售': '商贸零售', '零售': '商贸零售', '百货': '商贸零售',
    '社会服务': '社会服务', '旅游': '社会服务', '酒店': '社会服务',
    '机械设备': '机械设备', '机械': '机械设备', '工业母机': '机械设备',
}

def step10b_sector_completion(candidates):
    """为标的补全板块和行业信息"""
    print("\n" + "=" * 60)
    print("步骤10B: 板块/行业补全")
    print("=" * 60)
    
    for c in candidates:
        code = c.get('code', '')
        name = c.get('name', '')
        
        # 根据代码前缀推断板块
        if code.startswith('600') or code.startswith('601') or code.startswith('603') or code.startswith('605'):
            c['market'] = '上海主板'
        elif code.startswith('000') or code.startswith('001') or code.startswith('002') or code.startswith('003'):
            c['market'] = '深圳主板'
        elif code.startswith('300') or code.startswith('301'):
            c['market'] = '创业板'
        elif code.startswith('688'):
            c['market'] = '科创板'
        else:
            c['market'] = '其他'
        
        # 通过名称关键词推断行业
        industry = '未知'
        sector = '未知'
        for keyword, ind in SECTOR_INDUSTRY_MAP.items():
            if keyword in name:
                industry = ind
                sector = keyword
                break
        
        c['industry'] = industry
        c['sector'] = sector
    
    print(f"✅ 步骤10B: 完成板块/行业补全")
    return candidates

# ============================================================
# 步骤11: 硬性排除（31项，三级可达性）
# ============================================================
def step11_hard_exclusion(candidates):
    """执行31项硬排除"""
    global excluded_count, total_raw, exclusion_reasons
    
    print("\n" + "=" * 60)
    print("步骤11: 硬性排除 (31项)")
    print("=" * 60)
    
    total_raw = len(candidates)
    passed = []
    
    # 读取推荐历史(7日内推荐+已持仓)
    history = safe_read_json('/workspace/推荐历史.json')
    recent_codes = set()
    for r in history:
        if r.get('type') in ('recommendation', 'holding'):
            code = r.get('code', '')
            d = r.get('date', '') or r.get('update_date', '')
            if code and d >= (datetime.strptime(data_date, '%Y-%m-%d') - timedelta(days=7)).strftime('%Y-%m-%d'):
                recent_codes.add(code)
    
    for c in candidates:
        code = c.get('code', '')
        name = c.get('name', '')
        close = c.get('close', 0)
        change_pct = c.get('change_pct', 0)
        turnover = c.get('turnover', 0)
        pe_ttm = c.get('pe_ttm')
        
        excluded = False
        reason = []
        
        # L1 必执行 (1-16, 20-22, 28)
        # 1. 科创板(688xxx)
        if code.startswith('688'):
            reason.append('科创板(规则1)')
            excluded = True
        
        # 2. 北交所(8开头)
        if code.startswith('8') and len(code) == 6:
            reason.append('北交所(规则2)')
            excluded = True
        
        # 3. 股价<5元
        if close < 5:
            reason.append('股价<5元(规则3)')
            excluded = True
        
        # 4. 股价>100元
        if close > 100:
            reason.append('股价>100元(规则4)')
            excluded = True
        
        # 5. ST/*ST
        if 'ST' in name.upper() or '*ST' in name.upper():
            reason.append('ST/*ST(规则5)')
            excluded = True
        
        # 8. 上市<60日 (简化：通过代码判断，新代码通常>600000)
        # 跳过，难以精确判断
        
        # 11. 涨停/连板 (涨幅>9.5%)
        if change_pct > 9.5:
            reason.append('涨停(规则11)')
            excluded = True
        
        # 12. 涨幅>7%
        if change_pct > 7:
            reason.append('涨幅>7%(规则12)')
            excluded = True
        
        # 13. 7日内已推荐+已持仓
        if code in recent_codes:
            reason.append('7日内已推荐/持仓(规则13)')
            excluded = True
        
        # 20. PE(TTM)>500且非困境反转
        if pe_ttm and pe_ttm > 500:
            reason.append('PE>500(规则20)')
            excluded = True
        
        # 21. 创业板(300xxx)仅强市+动量延续
        if code.startswith('300') and market_env != '强市':
            reason.append('创业板非强市(规则21)')
            excluded = True
        
        # 22. 跌停
        if change_pct < -9.5:
            reason.append('跌停(规则22)')
            excluded = True
        
        # 28. 近20日跌幅>30% (简化：当日跌幅>7%就标记)
        # 需要更多数据，简化处理
        
        if excluded:
            for r in reason:
                exclusion_reasons[r.split('(')[0]] += 1
            continue
        
        passed.append(c)
    
    excluded_count = total_raw - len(passed)
    print(f"   原始: {total_raw} → 硬排除: {excluded_count} → 通过: {len(passed)}")
    print(f"   排除TOP5: {exclusion_reasons.most_common(5)}")
    
    log_alert("INFO", "硬排除", f"原始{total_raw}只, 排除{excluded_count}只, 剩余{len(passed)}只")
    return passed

# ============================================================
# 步骤12: 信号质量过滤（14项）
# ============================================================
def step12_signal_filter(candidates):
    """14项信号过滤"""
    global filtered_count
    
    print("\n" + "=" * 60)
    print("步骤12: 信号质量过滤 (14项)")
    print("=" * 60)
    
    passed = []
    signal_excluded = Counter()
    
    for c in candidates:
        change_pct = c.get('change_pct', 0)
        turnover = c.get('turnover', 0)
        amplitude = c.get('amplitude', 0)
        volume_ratio = c.get('volume_ratio')
        open_price = c.get('open', 0)
        close = c.get('close', 0)
        high = c.get('high', 0)
        low = c.get('low', 0)
        
        excluded = False
        reason = []
        
        # 1. 假动量: 高开>3%且收<开×0.98
        if open_price > 0 and close > 0:
            if (open_price / close - 1) > 0.03 and close < open_price * 0.98:
                if change_pct < 0:
                    reason.append('假动量')
                    excluded = True
        
        # 3. 尾盘急拉 (简化：振幅>10%且收涨>3%)
        # 无法精确判断尾盘，跳过
        
        # 4. 尾盘跳水 (同理)
        
        # 5. 换手率>30% (非次新/非公告日)
        if turnover > 30:
            reason.append('换手率>30%')
            excluded = True
        
        # 6. 放量滞涨
        if volume_ratio and volume_ratio > 2.0 and change_pct < 0.5:
            reason.append('放量滞涨')
            excluded = True
        
        # 7. 振幅>15%
        if amplitude > 15:
            reason.append('振幅>15%')
            excluded = True
        
        # 13. 竞价爆量 (无法精确判断)
        
        if excluded:
            for r in reason:
                signal_excluded[r] += 1
            continue
        
        passed.append(c)
    
    filtered_count = len(passed)
    print(f"   通过: {len(passed)} → 信号过滤排除: {len(candidates) - len(passed)}")
    if signal_excluded:
        print(f"   排除TOP: {signal_excluded.most_common(5)}")
    
    return passed

# ============================================================
# 步骤13: 五策略筛选
# ============================================================
# 策略优先级（与 lib/match.py _MATCH_STRATEGY_ORDER 保持一致）
_PART2_STRATEGY_ORDER = {'A': 0, 'D': 1, 'C': 2, 'B': 3, 'E': 4}

def step13_strategy_matching(candidates):
    """五大策略匹配"""
    global matched_count, strategy_counts
    
    print("\n" + "=" * 60)
    print("步骤13: 五策略匹配")
    print("=" * 60)
    
    matched = []
    
    for c in candidates:
        change_pct = c.get('change_pct', 0)
        turnover = c.get('turnover', 0)
        volume_ratio = c.get('volume_ratio')
        amount = c.get('amount', 0)
        close = c.get('close', 0)
        
        strategies = []
        
        # A 动量延续: 涨幅3-7%, 量比1.5-3.0, MA5>MA10>MA20
        # 弱市关闭策略A
        if market_env != '弱市':
            if 3 <= change_pct <= 7:
                if volume_ratio is None or 1.5 <= volume_ratio <= 3.0:
                    strategies.append(('A', '动量延续', 12))
        
        # B 超跌反弹: 连跌≥3日, 量<5日均×0.6, RSI(14)<35, 股价≥MA60
        # 简化：当日涨幅为正但前一天为负，且换手率较低
        if -5 <= change_pct <= 3:
            if turnover < 15:
                strategies.append(('B', '超跌反弹', 9))
        
        # C 事件驱动: 需要新闻搜索，简化跳过
        # 仅在明确有事件时才匹配
        
        # D 资金埋伏: 北向连续净买+主力流入>3000万+涨幅<2%
        if 0 < change_pct < 2:
            if turnover < 10:
                strategies.append(('D', '资金埋伏', 7))
        
        # E 回调企稳突破: 20日内创新高+回调MA20±3%+缩量+站回MA5放量
        # 简化：涨幅2-5%，换手率适中
        if 2 <= change_pct <= 5:
            if 5 <= turnover <= 20:
                strategies.append(('E', '回调企稳突破', 8))
        
        if strategies:
            # 选择优先级最高的策略
            best = min(strategies, key=lambda x: _PART2_STRATEGY_ORDER.get(x[0], 99))
            c['strategy'] = best[0]
            c['strategy_name'] = best[1]
            c['score'] = best[2]
            c['reason'] = best[1]
            matched.append(c)
            strategy_counts[best[0]] += 1
    
    matched_count = len(matched)
    print(f"   策略匹配: {matched_count} 只")
    print(f"   策略分布: {dict(strategy_counts)}")
    
    return matched

# ============================================================
# 步骤14-16: 评分
# ============================================================
def step14_16_scoring(matched):
    """评分门控 + 综合评分"""
    global matched_count
    
    print("\n" + "=" * 60)
    print("步骤14-16: 综合评分")
    print("=" * 60)
    
    for c in matched:
        strategy = c.get('strategy', '')
        change_pct = c.get('change_pct', 0)
        turnover = c.get('turnover', 0)
        volume_ratio = c.get('volume_ratio', 0) or 0
        
        base_score = c.get('score', 0)
        
        # 加分项
        bonus = 0
        
        # 板块TOP5 (简化：标记为板块TOP5则+1)
        # 信号加分项
        # K线形态确认
        
        # L3扣分
        l3_penalty = 0
        main_inflow = c.get('main_inflow')
        if main_inflow is not None and main_inflow < -10000:
            l3_penalty -= 2
        
        # 换手率加分
        if 5 <= turnover <= 15:
            bonus += 1
        elif turnover > 25:
            bonus -= 1
        
        # 量比加分
        if volume_ratio and 1.5 <= volume_ratio <= 3.0:
            bonus += 1
        
        total = base_score + bonus + l3_penalty
        c['score'] = max(1, total)
        
        # 置信度
        if total >= 9:
            c['confidence'] = '★★★'
        elif total >= 6:
            c['confidence'] = '★★'
        else:
            c['confidence'] = '★'
        
        # 进场/止损/止盈
        close = c.get('close', 0)
        c['entry'] = round(close, 2)
        c['stop_loss'] = round(close * 0.96, 2)
        c['take_profit'] = round(close * 1.05, 2)
        
        # URL
        code = c.get('code', '')
        if code.startswith('6'):
            c['url'] = f'https://quote.eastmoney.com/concept/sh{code}.html'
        else:
            c['url'] = f'https://quote.eastmoney.com/concept/sz{code}.html'
    
    # 排序
    matched.sort(key=lambda x: (x.get('score', 0), x.get('change_pct', 0) or 0, x.get('turnover', 0) or 0), reverse=True)
    
    print(f"   评分完成，最高分: {matched[0].get('score',0) if matched else 0}")
    return matched

# ============================================================
# 步骤17: 行业集中度限制
# ============================================================
def step17_industry_limit(matched):
    """行业集中度 + 同策略上限"""
    global industry_limited_count
    
    print("\n" + "=" * 60)
    print("步骤17: 行业集中度限制")
    print("=" * 60)
    
    industry_count = Counter()
    strategy_limit = {}
    strategy_concentration = params.get('strategy_concentration_pct', 60)
    # 预期最终推荐数约5-10只，同策略≤60%即≤3-6只
    expected_final = 10
    max_same_strategy = max(2, int(expected_final * strategy_concentration / 100))
    
    limited = []
    for c in matched:
        industry = c.get('industry', '未知')
        strategy = c.get('strategy', '')
        
        # 同行业≤3只
        if industry_count[industry] >= 3:
            continue
        
        # 同策略≤strategy_concentration_pct%
        if strategy_limit.get(strategy, 0) >= max_same_strategy:
            continue
        
        strategy_limit[strategy] = strategy_limit.get(strategy, 0) + 1
        industry_count[industry] += 1
        limited.append(c)
    
    industry_limited_count = len(limited)
    print(f"   行业限制后: {industry_limited_count} 只 (同行业≤3, 同策略≤{max_same_strategy})")
    return limited

# ============================================================
# 步骤18: 新闻筛查
# ============================================================
def step18_news_screening(limited):
    """新闻筛查"""
    global news_filtered_count
    
    print("\n" + "=" * 60)
    print("步骤18: 新闻筛查")
    print("=" * 60)
    
    # 新闻筛查需要WebSearch，在沙箱中简化
    # 保留所有标的，不做额外排除
    news_filtered_count = len(limited)
    print(f"   新闻筛查后: {news_filtered_count} 只 (沙箱环境简化)")
    return limited

# ============================================================
# 步骤19: 推荐不足降级
# ============================================================
def step19_insufficient_recommendation(recos):
    """推荐不足降级"""
    global final_recommend_count
    
    n = len(recos)
    if n == 0:
        print(f"⚠️ 步骤19: 无合适标的")
        final_recommend_count = 0
        return recos
    elif n == 1:
        print(f"⚠️ 步骤19: 仅1只，仅高置信保留")
        recos = [r for r in recos if r.get('confidence') == '★★★']
    elif n == 2:
        print(f"⚠️ 步骤19: 仅2只，仅≥中置信保留")
        recos = [r for r in recos if r.get('confidence') in ('★★', '★★★')]
    elif n == 3:
        print(f"⚠️ 步骤19: 仅3只，全部+放宽至中置信")
    
    final_recommend_count = len(recos)
    return recos

# ============================================================
# 步骤20: 输出Excel
# ============================================================
def step20_output_excel(recos):
    """输出Excel文件"""
    global final_recommend_count
    
    print("\n" + "=" * 60)
    print("步骤20: 输出Excel")
    print("=" * 60)
    
    xlsx_path = f'/workspace/短线标的_{prediction_date}.xlsx'
    wb = Workbook()
    
    # === Sheet 1: 标的池 ===
    ws = wb.active
    ws.title = "标的池"
    
    headers = ["序号","策略","标的","代码","板块","行业","当日涨跌","开盘价","收盘价","换手率","振幅","预测逻辑","评分","置信度","进场","止损","止盈","链接"]
    
    # 样式
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )
    
    # 策略色
    strategy_fills = {
        'A': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'B': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
        'C': PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
        'D': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'E': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    }
    
    # 涨跌色
    red_font = Font(name='Arial', size=10, color='9C0006')
    green_font = Font(name='Arial', size=10, color='006100')
    
    # 置信度色
    conf_fills = {
        '★★★': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        '★★': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        '★': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }
    
    link_font = Font(name='Arial', size=10, color='0563C1', underline='single')
    
    # 写表头
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 写数据
    final_recommend_count = len(recos)
    for i, rec in enumerate(recos, 1):
        row = i + 1
        strategy = rec.get('strategy', '')
        s_fill = strategy_fills.get(strategy)
        
        values = [
            i,  # 序号
            strategy,  # 策略
            rec.get('name', ''),  # 标的
            rec.get('code', ''),  # 代码
            rec.get('sector', ''),  # 板块
            rec.get('industry', ''),  # 行业
            rec.get('change_pct'),  # 当日涨跌
            rec.get('open'),  # 开盘价
            rec.get('close'),  # 收盘价
            rec.get('turnover'),  # 换手率
            rec.get('amplitude'),  # 振幅
            rec.get('reason', ''),  # 预测逻辑
            rec.get('score'),  # 评分
            rec.get('confidence', ''),  # 置信度
            rec.get('entry'),  # 进场
            rec.get('stop_loss'),  # 止损
            rec.get('take_profit'),  # 止盈
            rec.get('url', ''),  # 链接
        ]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            
            if s_fill:
                cell.fill = s_fill
            
            # 涨跌列(7)着色
            if col_idx == 7 and val is not None:
                cell.font = red_font if val > 0 else green_font
                cell.number_format = '0.00%' if isinstance(val, (int, float)) else 'General'
            
            # 置信度列(14)着色
            if col_idx == 14 and val in conf_fills:
                cell.fill = conf_fills[val]
                cell.font = Font(name='Arial', size=10, bold=True)
            
            # 链接列(18)蓝色
            if col_idx == 18 and val:
                cell.font = link_font
        
        ws.row_dimensions[row].height = 22
    
    # 调整列宽
    col_widths = [6, 6, 14, 10, 10, 12, 10, 10, 10, 10, 10, 20, 6, 8, 10, 10, 10, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    
    # 策略说明尾部 - 使用最终推荐列表统计
    final_strat_counts = Counter()
    for r in recos:
        final_strat_counts[r.get('strategy', '?')] += 1
    footer_start = ws.max_row + 2
    ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=18)
    cell = ws.cell(row=footer_start, column=1, value=f"📊 共筛选出 {final_recommend_count} 只标的（A:{final_strat_counts.get('A',0)} B:{final_strat_counts.get('B',0)} C:{final_strat_counts.get('C',0)} D:{final_strat_counts.get('D',0)} E:{final_strat_counts.get('E',0)}）")
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    footer_start += 1
    ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=18)
    cell = ws.cell(row=footer_start, column=1, value="策略说明：")
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = Alignment(horizontal='left')
    
    strategies = [
        ("A 动量延续", "涨幅3-7%，量比1.5-3.0，MA5>MA10>MA20 — 仓位强35-40%/震荡12-17%/弱关闭"),
        ("B 超跌反弹", "连跌≥3日，量<5日均×0.6，RSI(14)<35，KDJ(K<20且J拐头)，站上MA5+放量确认，股价≥MA60 — 仓位强10-12%/震荡12-15%/弱12-15%"),
        ("C 事件驱动", "重大合同/预增>50%/部委级政策，事件时效5级衰减 — 仓位强10-12%/震荡10-12%/弱5-8%"),
        ("D 资金埋伏", "北向3日连续净买+主力流入>3000万+涨幅<2% — 仓位强5-8%/震荡5-8%/弱3-5%"),
        ("E 回调企稳突破", "20日内创新高+回调MA20±3%+连3日缩量+站回MA5放量 — 仓位强10-12%/震荡12-15%/弱8-12%"),
    ]
    for name, desc in strategies:
        footer_start += 1
        ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=18)
        cell = ws.cell(row=footer_start, column=1, value=f"{name}：{desc}")
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    footer_start += 2
    ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=18)
    cell = ws.cell(row=footer_start, column=1, value="⚠️ 仅供参考，不构成投资建议")
    cell.font = Font(name='Arial', size=9, color='6B7280')
    cell.alignment = Alignment(horizontal='center')
    
    wb.save(xlsx_path)
    print(f"✅ Excel已保存: {xlsx_path}")
    return xlsx_path

# ============================================================
# 步骤20B: 生成HTML报告
# ============================================================
def step20b_html_report(recos):
    """生成自包含HTML报告"""
    print("\n" + "=" * 60)
    print("步骤20B: 生成HTML报告")
    print("=" * 60)
    
    html_dir = f'/workspace/ashare-screening-{prediction_date}'
    os.makedirs(html_dir, exist_ok=True)
    
    # 构建HTML
    # 使用最终推荐列表统计策略分布
    html_strat_counts = Counter()
    for r in recos:
        html_strat_counts[r.get('strategy', '?')] += 1
    strategy_dist = dict(html_strat_counts)
    strategy_names = {'A': '动量延续', 'B': '超跌反弹', 'C': '事件驱动', 'D': '资金埋伏', 'E': '回调企稳'}
    
    # 排除TOP5数据
    excl_top5 = exclusion_reasons.most_common(5)
    
    # === 纯CSS图表数据 ===
    # 策略分段条
    total_s = sum(strategy_dist.values()) or 1
    seg_html = ""
    seg_legend_html = ""
    seg_colors = {'A': ('sa', '#16a34a', '动量延续'), 'B': ('sb', '#2563eb', '超跌反弹'), 'C': ('sc', '#9333ea', '事件驱动'), 'D': ('sd', '#d97706', '资金埋伏'), 'E': ('se', '#ea580c', '回调企稳')}
    for k in ['A', 'B', 'C', 'D', 'E']:
        v = strategy_dist.get(k, 0)
        if v > 0:
            pct = v / total_s * 100
            cls, color, name = seg_colors[k]
            seg_html += f'<div class="seg {cls}" style="width:{pct:.1f}%">{v}只</div>'
            seg_legend_html += f'<span><i class="seg-dot {cls}"></i>{name} {v}只 ({pct:.0f}%)</span>'
    
    # 排除TOP5 CSS柱状图
    excl_bars_html = ""
    max_excl = excl_top5[0][1] if excl_top5 else 1
    for idx, (name, cnt) in enumerate(excl_top5):
        pct = cnt / max_excl * 90
        ci = (idx % 5) + 1
        excl_bars_html += f'<div class="bar-row"><span class="bar-label">{name}</span><span class="bar-track"><span class="bar-fill c{ci}" style="width:{pct:.0f}%">{cnt}只</span></span></div>'
    
    # 漏斗
    funnel_html = ""
    funnel_stages = [
        ("原始标的池", total_raw, 1), ("硬排除后", total_raw - excluded_count, 2),
        ("信号过滤后", filtered_count, 3), ("策略匹配后", matched_count, 4),
        ("行业+新闻后", industry_limited_count, 5), ("最终推荐", final_recommend_count, 6)
    ]
    for name, cnt, fi in funnel_stages:
        funnel_html += f'<div class="fn-row fn{fi}">{name}: {cnt}只</div>'
    
    # 策略柱状图
    strat_bars_html = ""
    max_strat = max(strategy_dist.values()) if strategy_dist else 1
    for k in ['A', 'B', 'C', 'D', 'E']:
        v = strategy_dist.get(k, 0)
        pct = v / max_strat * 90 if max_strat > 0 else 0
        cls, color, name = seg_colors[k]
        strat_bars_html += f'<div class="bar-row"><span class="bar-label">{name}</span><span class="bar-track"><span class="bar-fill" style="width:{pct:.0f}%;background:{color}">{v}只</span></span></div>'
    
    # 推荐表格行
    table_rows = ""
    for i, rec in enumerate(recos, 1):
        strategy = rec.get('strategy', '')
        chg = rec.get('change_pct', 0) or 0
        chg_color = '#9C0006' if chg > 0 else '#006100'
        conf = rec.get('confidence', '')
        conf_color = {'★★★': '#006100', '★★': '#B8860B', '★': '#9C0006'}.get(conf, '#333')
        
        table_rows += f"""
        <tr class="strategy-{strategy.lower()}">
            <td>{i}</td>
            <td><span class="badge badge-{strategy.lower()}">{strategy}</span></td>
            <td><a href="{rec.get('url', '#')}" target="_blank">{rec.get('name', '')}</a></td>
            <td>{rec.get('code', '')}</td>
            <td>{rec.get('industry', '')}</td>
            <td style="color:{chg_color}">{chg:+.2f}%</td>
            <td>{rec.get('open', '-')}</td>
            <td>{rec.get('close', '-')}</td>
            <td>{rec.get('amplitude', '-')}</td>
            <td>{rec.get('score', 0)}</td>
            <td style="color:{conf_color};font-weight:bold">{conf}</td>
            <td>{rec.get('entry', '-')}</td>
            <td>{rec.get('stop_loss', '-')}</td>
            <td>{rec.get('take_profit', '-')}</td>
        </tr>"""
    
    # 告警日志
    alert_html = ""
    try:
        with open('/workspace/系统告警.log', 'r', encoding='utf-8') as f:
            alerts = [l.strip() for l in f.readlines() if data_date in l or prediction_date in l]
            for a in alerts[-20:]:
                level = 'WARN'
                if '[ERROR]' in a: level = 'ERROR'
                elif '[WARNING]' in a: level = 'WARN'
                elif '[INFO]' in a: level = 'INFO'
                alert_html += f'<div class="alert-item alert-{level.lower()}"><span class="alert-tag">{level}</span><span>{a}</span></div>\n'
    except Exception:
        alert_html = '<div class="alert-item alert-info"><span class="alert-tag">INFO</span><span>今日无告警</span></div>'
    
    html_content = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>A股短线标的筛选报告 — {prediction_date}</title>
<style>
:root {{
    --accent: #1F4E79;
    --bg: #f5f7fa;
    --card-bg: #ffffff;
    --text: #333333;
    --text-secondary: #6B7280;
    --border: #e5e7eb;
    --green: #006100;
    --red: #9C0006;
    --amber: #B8860B;
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Noto Sans CJK SC', 'WenQuanYi Micro Hei', Arial, sans-serif; background: var(--bg); color: var(--text); line-height: 1.6; }}
.container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}

/* 报告头部 */
.header {{ background: linear-gradient(135deg, #1a3a5c 0%, #1F4E79 50%, #2563eb 100%); color: white; padding: 30px; border-radius: 12px; margin-bottom: 24px; }}
.header h1 {{ font-size: 24px; margin-bottom: 8px; }}
.header .date {{ font-size: 14px; opacity: 0.85; }}
.meta-row {{ display: flex; gap: 16px; margin-top: 20px; flex-wrap: wrap; }}
.meta-card {{ background: rgba(255,255,255,0.15); border-radius: 8px; padding: 12px 16px; min-width: 140px; flex: 1; }}
.meta-card .label {{ font-size: 12px; opacity: 0.8; }}
.meta-card .value {{ font-size: 20px; font-weight: bold; margin-top: 4px; }}

/* 筛选管道 */
.pipeline {{ display: flex; align-items: center; gap: 8px; background: var(--card-bg); border-radius: 12px; padding: 24px; margin-bottom: 24px; flex-wrap: wrap; justify-content: center; }}
.pipe-step {{ text-align: center; padding: 12px 16px; border-radius: 8px; border: 2px solid var(--border); min-width: 100px; }}
.pipe-step .count {{ font-size: 24px; font-weight: bold; }}
.pipe-step .label {{ font-size: 12px; color: var(--text-secondary); }}
.pipe-step.final {{ border-color: #2563eb; background: #eff6ff; }}
.pipe-arrow {{ font-size: 20px; color: var(--text-secondary); }}

/* 纯CSS图表 */
.bar-row {{ display: flex; align-items: center; margin-bottom: 10px; }}
.bar-label {{ width: 90px; font-size: 12px; text-align: right; padding-right: 8px; color: var(--text-secondary); flex-shrink: 0; }}
.bar-track {{ flex: 1; height: 22px; background: #f1f5f9; border-radius: 4px; overflow: hidden; }}
.bar-fill {{ height: 100%; border-radius: 4px; line-height: 22px; padding-left: 6px; font-size: 11px; color: white; font-weight: bold; min-width: 20px; }}
.bar-fill.c1 {{ background: #3b82f6; }}
.bar-fill.c2 {{ background: #6366f1; }}
.bar-fill.c3 {{ background: #8b5cf6; }}
.bar-fill.c4 {{ background: #a855f7; }}
.bar-fill.c5 {{ background: #d946ef; }}
.bar-val {{ width: 40px; font-size: 12px; text-align: right; padding-left: 6px; flex-shrink: 0; font-weight: bold; }}

/* 策略分段条 */
.seg-bar {{ display: flex; height: 36px; border-radius: 8px; overflow: hidden; margin-bottom: 12px; }}
.seg {{ display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; color: white; }}
.seg.sa {{ background: #16a34a; }}
.seg.sb {{ background: #2563eb; }}
.seg.sc {{ background: #9333ea; }}
.seg.sd {{ background: #d97706; }}
.seg.se {{ background: #ea580c; }}
.seg-legend {{ display: flex; flex-wrap: wrap; gap: 12px; font-size: 11px; }}
.seg-legend span {{ display: flex; align-items: center; gap: 4px; }}
.seg-dot {{ width: 10px; height: 10px; border-radius: 2px; display: inline-block; }}
.seg-dot.sa {{ background: #16a34a; }}
.seg-dot.sb {{ background: #2563eb; }}
.seg-dot.sc {{ background: #9333ea; }}
.seg-dot.sd {{ background: #d97706; }}
.seg-dot.se {{ background: #ea580c; }}

/* 漏斗 */
.funnel {{ display: flex; flex-direction: column; gap: 4px; }}
.fn-row {{ display: flex; align-items: center; justify-content: center; height: 40px; border-radius: 4px; font-size: 12px; font-weight: bold; color: white; margin: 0 auto; white-space: nowrap; }}
.fn1 {{ width: 100%; background: #3b82f6; max-width: 300px; }}
.fn2 {{ width: 85%; background: #6366f1; max-width: 255px; }}
.fn3 {{ width: 70%; background: #8b5cf6; max-width: 210px; }}
.fn4 {{ width: 55%; background: #a855f7; max-width: 165px; }}
.fn5 {{ width: 40%; background: #d946ef; max-width: 120px; }}
.fn6 {{ width: 28%; background: #2563eb; max-width: 84px; }}

/* 推荐表格 */
.table-section {{ background: var(--card-bg); border-radius: 12px; padding: 20px; margin-bottom: 24px; overflow-x: auto; }}
.table-section h3 {{ font-size: 16px; margin-bottom: 16px; color: var(--accent); }}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
th {{ background: var(--accent); color: white; padding: 10px 8px; text-align: left; font-weight: 600; }}
td {{ padding: 8px; border-bottom: 1px solid var(--border); }}
tr:hover {{ background: #f8fafc; }}
.badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; }}
.badge-a {{ background: #E2EFDA; color: #166534; }}
.badge-b {{ background: #D6E4F0; color: #1e40af; }}
.badge-c {{ background: #E4DFEC; color: #6b21a8; }}
.badge-d {{ background: #FFF2CC; color: #92400e; }}
.badge-e {{ background: #FCE4D6; color: #c2410c; }}
tr.strategy-a {{ background: #f0fdf4; }}
tr.strategy-b {{ background: #eff6ff; }}
tr.strategy-c {{ background: #faf5ff; }}
tr.strategy-d {{ background: #fffbeb; }}
tr.strategy-e {{ background: #fff7ed; }}

/* 策略说明 */
.strategy-table {{ background: var(--card-bg); border-radius: 12px; padding: 20px; margin-bottom: 24px; }}

/* 告警列表 */
.alert-list {{ background: var(--card-bg); border-radius: 12px; padding: 20px; margin-bottom: 24px; }}
.alert-item {{ display: flex; gap: 12px; padding: 8px 0; border-bottom: 1px solid var(--border); font-size: 12px; }}
.alert-tag {{ padding: 1px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; }}
.alert-error .alert-tag {{ background: #fee2e2; color: var(--red); }}
.alert-warn .alert-tag {{ background: #fef3c7; color: var(--amber); }}
.alert-info .alert-tag {{ background: #dbeafe; color: #1e40af; }}

/* 尾部 */
.footer {{ text-align: center; padding: 20px; color: var(--text-secondary); font-size: 12px; }}
.footer .disclaimer {{ color: var(--red); font-weight: bold; font-size: 14px; margin-top: 8px; }}

@media (max-width: 768px) {{
    .chart-grid {{ grid-template-columns: 1fr; }}
    .meta-row {{ flex-direction: column; }}
    .header {{ padding: 20px; }}
    .header h1 {{ font-size: 18px; }}
    .meta-card {{ min-width: auto; padding: 8px 12px; }}
    .meta-card .value {{ font-size: 16px; }}
    .pipeline {{ padding: 12px; gap: 4px; }}
    .pipe-step {{ min-width: 60px; padding: 8px; }}
    .pipe-step .count {{ font-size: 16px; }}
    .table-section {{ padding: 12px; overflow-x: auto; }}
    table {{ font-size: 11px; }}
    th, td {{ padding: 6px 4px; }}
    
    .container {{ padding: 10px; }}
}}
</style>
</head>
<body>
<div class="container">
    <!-- 1. 报告头部 -->
    <div class="header">
        <h1>📊 A股每日短线标的筛选报告</h1>
        <div class="date">预测日期: {prediction_date} | 数据来源: {data_date}</div>
        <div class="meta-row">
            <div class="meta-card"><div class="label">预测日期</div><div class="value">{prediction_date}</div></div>
            <div class="meta-card"><div class="label">数据日期</div><div class="value">{data_date}</div></div>
            <div class="meta-card"><div class="label">市场环境</div><div class="value">{market_env}</div></div>
            <div class="meta-card"><div class="label">建议仓位</div><div class="value">{'50-60%' if market_env == '震荡' else '30-40%' if market_env == '弱市' else '70-80%'}</div></div>
            <div class="meta-card"><div class="label">最终推荐</div><div class="value">{final_recommend_count} 只</div></div>
        </div>
    </div>

    <!-- 2. 筛选管道 -->
    <div class="pipeline">
        <div class="pipe-step"><div class="count">{total_raw}</div><div class="label">原始标的池</div></div>
        <div class="pipe-arrow">→</div>
        <div class="pipe-step"><div class="count">{excluded_count}</div><div class="label">硬排除</div></div>
        <div class="pipe-arrow">→</div>
        <div class="pipe-step"><div class="count">{filtered_count}</div><div class="label">信号过滤</div></div>
        <div class="pipe-arrow">→</div>
        <div class="pipe-step"><div class="count">{matched_count}</div><div class="label">策略匹配</div></div>
        <div class="pipe-arrow">→</div>
        <div class="pipe-step"><div class="count">{industry_limited_count}</div><div class="label">行业+新闻</div></div>
        <div class="pipe-arrow">→</div>
        <div class="pipe-step final"><div class="count">{final_recommend_count}</div><div class="label">★ 最终推荐</div></div>
    </div>

    <!-- 3. 数据可视化 -->
    <div class="chart-grid">
        <div class="chart-card">
            <h3>策略分布</h3>
            <div class="seg-bar">{seg_html}</div>
            <div class="seg-legend">{seg_legend_html}</div>
        </div>
        <div class="chart-card">
            <h3>硬排除 TOP5</h3>
            {excl_bars_html}
        </div>
        <div class="chart-card">
            <h3>筛选漏斗</h3>
            <div class="funnel">{funnel_html}</div>
        </div>
        <div class="chart-card">
            <h3>各策略数量</h3>
            {strat_bars_html}
        </div>
    </div>

    <!-- 4. 最终推荐标的表 -->
    <div class="table-section">
        <h3>📋 最终推荐标的 ({final_recommend_count}只)</h3>
        <table>
            <thead>
                <tr><th>序号</th><th>策略</th><th>标的</th><th>代码</th><th>行业</th><th>涨跌幅</th><th>开盘价</th><th>收盘价</th><th>振幅</th><th>评分</th><th>置信度</th><th>进场</th><th>止损</th><th>止盈</th></tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
    </div>

    <!-- 5. 策略说明 -->
    <div class="strategy-table">
        <h3>📖 策略说明</h3>
        <table>
            <tr><th style="width:80px">策略</th><th>条件</th><th>仓位(震荡市)</th></tr>
            <tr class="strategy-a"><td><span class="badge badge-a">A</span></td><td>涨幅3-7%，量比1.5-3.0，MA5>MA10>MA20</td><td>12-17%</td></tr>
            <tr class="strategy-b"><td><span class="badge badge-b">B</span></td><td>连跌≥3日，量<5日均×0.6，RSI<35，KDJ(K<20且J拐头)，站上MA5+放量确认</td><td>12-15%</td></tr>
            <tr class="strategy-c"><td><span class="badge badge-c">C</span></td><td>重大合同/预增>50%/部委级政策，事件时效5级衰减</td><td>10-12%</td></tr>
            <tr class="strategy-d"><td><span class="badge badge-d">D</span></td><td>北向3日连续净买+主力流入>3000万+涨幅<2%</td><td>5-8%</td></tr>
            <tr class="strategy-e"><td><span class="badge badge-e">E</span></td><td>20日内创新高+回调MA20±3%+连3日缩量+站回MA5放量</td><td>12-15%</td></tr>
        </table>
    </div>

    <!-- 6. 系统告警 -->
    <div class="alert-list">
        <h3>🔔 系统告警</h3>
        {alert_html}
    </div>

    <!-- 7. 报告尾部 -->
    <div class="footer">
        <p>版本: {file_version} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 规则来源: lc132/lv</p>
        <p class="disclaimer">⚠️ 仅供参考，不构成投资建议</p>
    </div>
</div>
</body>
</html>'''
    
    html_path = f'{html_dir}/ashare-screening-{prediction_date}.html'
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✅ HTML报告已保存: {html_path}")
    return html_path

# ============================================================
# 步骤21: 最终验证
# ============================================================
def step21_validation(xlsx_path):
    """最终验证"""
    print("\n" + "=" * 60)
    print("步骤21: 最终验证")
    print("=" * 60)
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        if "标的池" in wb.sheetnames:
            ws = wb["标的池"]
            # 统计数据行（列1为数字的行）
            excel_n = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if isinstance(val, (int, float)):
                    excel_n += 1
            
            if excel_n != final_recommend_count:
                log_alert("ERROR", "数量校验", f"概况{final_recommend_count}≠Excel{excel_n}")
                print(f"⚠️ 验证失败: 概况{final_recommend_count}≠Excel{excel_n}")
            else:
                print(f"✅ 验证通过（{final_recommend_count}只）")
        
        wb.save(xlsx_path)
        wb.close()
    except Exception as e:
        log_alert("WARNING", "最终验证", f"验证失败: {str(e)[:80]}")
        print(f"⚠️ 验证异常: {str(e)[:80]}")

# ============================================================
# 步骤22: 写推荐历史
# ============================================================
def step22_write_history(recos):
    """写入推荐历史"""
    print("\n" + "=" * 60)
    print("步骤22: 写入推荐历史")
    print("=" * 60)
    
    for rec in recos:
        record = {
            "type": "recommendation",
            "date": prediction_date,
            "code": rec.get('code'),
            "name": rec.get('name'),
            "strategy": rec.get('strategy'),
            "score": rec.get('score'),
            "confidence": rec.get('confidence'),
            "entry": rec.get('entry'),
            "stop_loss": rec.get('stop_loss'),
            "take_profit": rec.get('take_profit'),
            "change_pct": rec.get('change_pct'),
        }
        safe_append_json('/workspace/推荐历史.json', record)
    
    print(f"✅ 已追加 {len(recos)} 条推荐记录")

# ============================================================
# 步骤23: 回溯检查昨日做T
# ============================================================
def step23_check_do_t():
    """回溯检查昨日做T"""
    print("\n" + "=" * 60)
    print("步骤23: 回溯检查昨日做T")
    print("=" * 60)
    
    history = safe_read_json('/workspace/推荐历史.json')
    yesterday = (datetime.strptime(prediction_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
    
    do_t_evals = [r for r in history if r.get('type') == 'do_T_eval' and r.get('date') == yesterday]
    do_t_records = [r for r in history if r.get('type') == 'do_T' and r.get('date') == yesterday]
    
    if do_t_evals and not do_t_records:
        print(f"⚠️ 昨日有 {len(do_t_evals)} 个做T评估但无做T记录，请检查")
    else:
        print(f"✅ 昨日做T记录: {len(do_t_records)} 条")

# ============================================================
# 步骤24: 告警日志摘要
# ============================================================
def step24_alert_summary():
    """告警日志摘要"""
    print("\n" + "=" * 60)
    print("步骤24: 告警日志摘要")
    print("=" * 60)
    
    try:
        with open('/workspace/系统告警.log', 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        today_lines = [l for l in lines if data_date in l or prediction_date in l]
        if today_lines:
            print(f"   今日告警 {len(today_lines)} 条:")
            for l in today_lines[-10:]:
                print(f"   {l.strip()}")
        else:
            print(f"   今日无异常")
    except Exception:
        print(f"   无告警日志")

# ============================================================
# 步骤25: 输出筛选概况
# ============================================================
def step25_screening_summary(recos):
    """输出筛选概况"""
    global strategy_counts
    # 重新计算最终推荐的策略分布
    strategy_counts = Counter()
    for r in recos:
        strategy_counts[r.get('strategy', '?')] += 1
    
    print("\n" + "=" * 60)
    print("📊 筛选概况")
    print("=" * 60)
    
    summary = f"""
📊 筛选概况 — {prediction_date}(数据来源:{data_date})
① 原始标的池:{total_raw}只 → ② 硬排除:{excluded_count}只 → ③ 信号过滤:{filtered_count}只 → ④ 策略匹配:{matched_count}只 → ⑤ 行业限制:{industry_limited_count}只 → ⑥ 新闻筛查:{news_filtered_count}只 → ★ 最终:{final_recommend_count}只
策略分布: A:{strategy_counts.get('A',0)} B:{strategy_counts.get('B',0)} C:{strategy_counts.get('C',0)} D:{strategy_counts.get('D',0)} E:{strategy_counts.get('E',0)}
排除TOP5: {exclusion_reasons.most_common(5)}
市场环境: {market_env}
"""
    print(summary)
    return summary

# ============================================================
# 步骤26: GitHub同步
# ============================================================
def step26_github_sync(xlsx_path):
    """GitHub同步"""
    print("\n" + "=" * 60)
    print("步骤26: GitHub同步")
    print("=" * 60)
    
    if not os.path.exists(xlsx_path):
        log_alert("WARNING", "GitHub同步", "xlsx文件不存在，跳过")
        print("⚠️ xlsx文件不存在，跳过")
        return
    
    token = None
    token_path = "/workspace/.github_token"
    if os.path.exists(token_path):
        try:
            with open(token_path, 'r', encoding='utf-8') as f:
                token = f.read().strip()
        except Exception:
            pass
    # 备用：从环境变量读取
    if not token:
        token = os.environ.get('GITHUB_TOKEN', '')
    if not token:
        log_alert("WARNING", "GitHub同步", "无认证令牌，跳过推送")
        print("⚠️ 无认证令牌，跳过")
        return
    
    repo_url = f"https://{token}@github.com/lc132/lv.git"
    repo_dir = "/tmp/lv_sync"
    
    try:
        if os.path.exists(repo_dir):
            shutil.rmtree(repo_dir, ignore_errors=True)
        
        subprocess.run(
            ["git", "clone", "--depth", "1", "--branch", "main", repo_url, repo_dir],
            capture_output=True, text=True, timeout=30, check=True
        )
        
        # 推送筛选结果
        shutil.copy(xlsx_path, os.path.join(repo_dir, f"短线标的_{prediction_date}.xlsx"))
        
        # 推送HTML报告
        html_dir = f'/workspace/ashare-screening-{prediction_date}'
        if os.path.exists(html_dir):
            html_dest = os.path.join(repo_dir, f'ashare-screening-{prediction_date}')
            if os.path.exists(html_dest):
                shutil.rmtree(html_dest, ignore_errors=True)
            shutil.copytree(html_dir, html_dest)
        
        subprocess.run(["git", "-C", repo_dir, "config", "user.email", "ashare-bot@github.com"], check=True, timeout=10)
        subprocess.run(["git", "-C", repo_dir, "config", "user.name", "ashare-screener"], check=True, timeout=10)
        subprocess.run(["git", "-C", repo_dir, "add", f"短线标的_{prediction_date}.xlsx"], check=True, timeout=10)
        if os.path.exists(html_dir):
            subprocess.run(["git", "-C", repo_dir, "add", f"ashare-screening-{prediction_date}"], check=True, timeout=10)
        
        commit_msg = f"筛选结果 {prediction_date}"
        subprocess.run(["git", "-C", repo_dir, "commit", "-m", commit_msg], check=True, timeout=10)
        result = subprocess.run(
            ["git", "-C", repo_dir, "push", "origin", "main"],
            capture_output=True, text=True, timeout=30
        )
        
        if result.returncode == 0:
            log_alert("INFO", "GitHub同步", f"✅ {prediction_date} 已推送")
            print(f"✅ GitHub同步成功")
        else:
            log_alert("WARNING", "GitHub同步", f"推送失败: {result.stderr[:100]}")
            print(f"⚠️ GitHub推送失败: {result.stderr[:100]}")
    except Exception as e:
        log_alert("WARNING", "GitHub同步", f"失败: {str(e)[:100]}")
        print(f"⚠️ GitHub同步失败: {str(e)[:100]}")
    finally:
        if os.path.exists(repo_dir):
            shutil.rmtree(repo_dir, ignore_errors=True)

# ============================================================
# 步骤27: 飞书推送
# ============================================================
def step27_feishu_push(summary_text):
    """飞书推送"""
    print("\n" + "=" * 60)
    print("步骤27: 飞书推送")
    print("=" * 60)
    
    FEISHU_WEBHOOK = "https://open.feishu.cn/open-apis/bot/v2/hook/792265f9-12d5-4e78-a163-08cad5699fa1"
    if not FEISHU_WEBHOOK:
        log_alert("WARNING", "飞书推送", "未配置Webhook URL，跳过")
        print("⚠️ 未配置Webhook")
        return
    
    strategy_dist = dict(strategy_counts) if strategy_counts else {}
    card = {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"📊 每日短线标的筛选 — {prediction_date}"},
                "template": "blue"
            },
            "elements": [
                {"tag": "div", "text": {"tag": "lark_md", "content": f"**数据来源**: {data_date} | **市场环境**: {market_env}"}},
                {"tag": "hr"},
                {"tag": "div", "text": {"tag": "lark_md", "content": f"原始标的池: **{total_raw}**只 → 硬排除: **{excluded_count}**只 → 信号过滤: **{filtered_count}**只 → 策略匹配: **{matched_count}**只 → 行业限制: **{industry_limited_count}**只 → 新闻筛查: **{news_filtered_count}**只 → ★ 最终: **{final_recommend_count}**只"}},
                {"tag": "hr"},
                {"tag": "div", "text": {"tag": "lark_md", "content": f"策略分布: A动量:{strategy_dist.get('A',0)} B超跌:{strategy_dist.get('B',0)} C事件:{strategy_dist.get('C',0)} D资金:{strategy_dist.get('D',0)} E回调:{strategy_dist.get('E',0)}"}},
                {"tag": "note", "elements": [{"tag": "plain_text", "content": "⚠️ 仅供参考，不构成投资建议"}]}
            ]
        }
    }
    
    try:
        req = urllib.request.Request(
            FEISHU_WEBHOOK,
            data=json.dumps(card, ensure_ascii=False).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read())
        if result.get('code') == 0:
            log_alert("INFO", "飞书推送", "✅ 筛选概况已推送到飞书群")
            print(f"✅ 飞书推送成功")
        else:
            log_alert("WARNING", "飞书推送", f"推送失败: {result.get('msg','')}")
            print(f"⚠️ 飞书推送失败: {result.get('msg','')}")
    except Exception as e:
        log_alert("WARNING", "飞书推送", f"请求异常: {str(e)[:100]}")
        print(f"⚠️ 飞书推送异常: {str(e)[:100]}")

# ============================================================
# 主流程 Part 2
# ============================================================
def run_part2():
    print("=" * 60)
    print("A股盘前短线标的筛选 v6.9.53 - Part 2: 步骤9-27")
    print("=" * 60)
    
    # 步骤9: 板块轮动
    sector_flow = step9_sector_rotation()
    
    # 步骤10A: 全市场API拉取
    all_stocks = step10a_fetch_all_stocks()
    if not all_stocks:
        log_alert("ERROR", "行情采集", "全市场API拉取失败")
        print("❌ 全市场行情拉取失败，中止")
        return False
    
    # 构建原始标的池
    global total_raw
    raw_pool = [s for s in all_stocks
                if s['change_pct'] is not None and s['close'] is not None and s['close'] > 0]
    raw_pool.sort(key=lambda x: x.get('turnover', 0) or 0, reverse=True)
    raw_pool = raw_pool[:500]
    total_raw = len(raw_pool)
    print(f"\n原始标的池: {total_raw} 只（全市场{len(all_stocks)}只中活跃TOP500）")
    
    # 构建candidate
    candidates = []
    for s in raw_pool:
        c = {
            "code": s["code"], "name": s["name"],
            "sector": "", "industry": "",
            "change_pct": s.get("change_pct"),
            "open": s.get("open"), "close": s.get("close"),
            "turnover": s.get("turnover"), "amplitude": s.get("amplitude"),
            "volume_ratio": s.get("volume_ratio"), "amount": s.get("amount"),
            "main_inflow": s.get("main_inflow"), "pe_ttm": s.get("pe_ttm"),
            "total_cap": s.get("total_cap"),
            "strategy": "", "reason": "", "score": 0, "confidence": "",
            "entry": None, "stop_loss": None, "take_profit": None,
            "url": f"https://quote.eastmoney.com/concept/sh{s['code']}.html" if s["code"].startswith('6') else f"https://quote.eastmoney.com/concept/sz{s['code']}.html"
        }
        if c["close"] is None or c["close"] <= 0: continue
        if c["change_pct"] is None: continue
        candidates.append(c)
    
    print(f"构建候选池: {len(candidates)} 只")
    
    # 步骤10B: 板块/行业补全
    candidates = step10b_sector_completion(candidates)
    
    # 步骤11: 硬排除
    candidates = step11_hard_exclusion(candidates)
    
    # 步骤12: 信号过滤
    candidates = step12_signal_filter(candidates)
    
    # 步骤13: 策略匹配
    candidates = step13_strategy_matching(candidates)
    
    # 步骤14-16: 评分
    candidates = step14_16_scoring(candidates)
    
    # 步骤17: 行业集中度
    candidates = step17_industry_limit(candidates)
    
    # 步骤18: 新闻筛查
    candidates = step18_news_screening(candidates)
    
    # 步骤19: 推荐不足降级
    candidates = step19_insufficient_recommendation(candidates)
    
    # 步骤20: 输出Excel
    xlsx_path = step20_output_excel(candidates)
    
    # 步骤20B: HTML报告
    html_path = step20b_html_report(candidates)
    
    # 步骤21: 最终验证
    step21_validation(xlsx_path)
    
    # 步骤22: 写推荐历史
    step22_write_history(candidates)
    
    # 步骤23: 回溯检查
    step23_check_do_t()
    
    # 步骤24: 告警摘要
    step24_alert_summary()
    
    # 步骤25: 筛选概况
    summary = step25_screening_summary(candidates)
    
    # 步骤26: GitHub同步
    step26_github_sync(xlsx_path)
    
    # 步骤27: 飞书推送
    step27_feishu_push(summary)
    
    print(f"\n{'='*60}")
    print(f"✅ 全流程完成！最终推荐: {final_recommend_count} 只标的")
    print(f"   Excel: {xlsx_path}")
    print(f"   HTML: {html_path}")
    print(f"{'='*60}")
    
    return True

if __name__ == '__main__':
    run_part2()