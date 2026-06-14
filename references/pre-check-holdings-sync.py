from openpyxl import load_workbook
def sync_holding_prices_to_xlsx(holdings, path="/workspace/持仓跟踪.xlsx"):
    """将步骤4更新后的持仓价格写入持仓跟踪.xlsx"""
    try:
        wb = load_workbook(path)
        ws = wb["持仓明细"]
        # code → row mapping (skip header row)
        code_row = {}
        for row in range(2, ws.max_row + 1):
            raw_code = ws.cell(row=row, column=1).value
            if raw_code:
                code = str(raw_code).strip()
                if len(code) == 4:  # Excel可能丢失前导零
                    code = code.zfill(6)
                if code.isdigit() and len(code) == 6:
                    code_row[code] = row
        
        updated = 0
        for h in holdings:
            current = None
            try:
                raw_code = h.get("code")
                code = str(raw_code) if raw_code is not None else ""
                current = h.get("current")  # 防御性读取，缺失返回None
                if not code or code not in code_row:
                    if code:
                        log_alert("WARNING", "持仓跟踪同步", f"{code} 在xlsx中找不到")
                    continue
                if current is None:
                    log_alert("WARNING", "持仓跟踪同步", f"{code} 缺少current字段，跳过")
                    continue

                row = code_row[code]
                # 市值和盈亏额：优先取holding中的值，缺失则从xlsx读取成本/持仓量计算
                mv = h.get("market_value")
                pnl_amt = h.get("pnl_amount")
                if mv is None or pnl_amt is None:
                    cost = ws.cell(row=row, column=3).value   # 成本
                    shares = ws.cell(row=row, column=4).value  # 持仓量
                    if cost and shares and current:
                        mv = round(current * shares, 2)
                        pnl_amt = round((current - cost) * shares, 2)
                ws.cell(row=row, column=8).value = current              # 当前价 (H)
                if mv is not None:
                    ws.cell(row=row, column=9).value = mv               # 市值 (I)
                if pnl_amt is not None:
                    ws.cell(row=row, column=10).value = round(pnl_amt, 2)  # 盈亏额 (J)
                pnl_pct_val = h.get("pnl_pct")
                try:
                    pnl_pct_float = float(pnl_pct_val) if pnl_pct_val is not None else 0.0
                except (ValueError, TypeError):
                    pnl_pct_float = 0.0
                ws.cell(row=row, column=11).value = round(pnl_pct_float, 4)  # 盈亏率 (K)
                ws.cell(row=row, column=12).value = beijing_date  # 更新日期 (L)
                updated += 1
            except Exception as e:
                log_alert("WARNING", "持仓跟踪同步", f"单条记录异常(code={h.get('code', 'unknown')}): {str(e)[:80]}")
                continue
        if updated > 0:
            wb.save(path)
            log_alert("INFO", "持仓跟踪同步", f"已更新{updated}只持仓价格")
    except Exception as e:
        log_alert("WARNING", "持仓跟踪同步", f"失败: {str(e)[:100]}")
