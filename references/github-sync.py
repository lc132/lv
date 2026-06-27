import subprocess, os, shutil

md_path = f"/workspace/短线标的_{prediction_date}.md"
try:
    with open(md_path, 'r', encoding='utf-8') as _f:
        pass
except FileNotFoundError:
    log_alert("WARNING", "GitHub同步", "md文件不存在，跳过")
    return

# 读取认证令牌
token = None
token_path = "/workspace/.github_token"
try:
    with open(token_path, 'r', encoding='utf-8') as f:
        token = f.read().strip()
except (FileNotFoundError, PermissionError):
    pass
if not token:
    log_alert("WARNING", "GitHub同步", "无认证令牌，跳过推送")
    return

# === 推送前校验并同步筛选条件表格 ===
cond_xlsx = "/workspace/A股短线选股筛选条件.xlsx"
cond_synced = False  # 仅在版本不一致且成功同步后置为 True
xlsx_version = None  # 显式初始化，防止 NameError
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb_cond = load_workbook(cond_xlsx)
    ws1 = wb_cond['筛选条件概述']
    # 读取xlsx中已记录的版本号（第2行第2列）
    xlsx_version = ws1.cell(row=2, column=2).value
    if xlsx_version and str(xlsx_version) != str(file_version):
        log_alert("INFO", "筛选条件", f"版本不一致: xlsx={xlsx_version} ≠ 当前={file_version}，先同步")

        _cell_font = Font(name='Arial', size=10)
        _bold_font = Font(name='Arial', size=10, bold=True)
        _thin_border = Border(
            left=Side(style='thin', color='B0B0B0'),
            right=Side(style='thin', color='B0B0B0'),
            top=Side(style='thin', color='B0B0B0'),
            bottom=Side(style='thin', color='B0B0B0'),
        )

        # _wc 函数与步骤6中定义一致，提取为独立函数避免重复维护
        def _wc(ws, r, c, v, font=_cell_font):
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                    if not (r == mr.min_row and c == mr.min_col):
                        return
                    ws.unmerge_cells(str(mr))
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = font
            cell.border = _thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        _wc(ws1, 1, 1, f'A股短线选股筛选条件 — {file_version}', _bold_font)
        _wc(ws1, 2, 2, file_version)
        _wc(ws1, 2, 3, f'{beijing_date}更新')
        vr = ws1.max_row + 1
        _wc(ws1, vr, 1, file_version)
        _wc(ws1, vr, 2, beijing_date)
        _wc(ws1, vr, 3, 'GitHub推送前自动同步')
        if '关键纪律' in wb_cond.sheetnames:
            ws11 = wb_cond['关键纪律']
            _wc(ws11, 1, 1, f'关键纪律 — {file_version}', _bold_font)

        wb_cond.save(cond_xlsx)
        wb_cond.close()
        cond_synced = True  # 标记已同步，推送时一并上传
        log_alert("INFO", "筛选条件", f"筛选条件.xlsx 已同步至 {file_version}")
    else:
        log_alert("INFO", "筛选条件", f"版本一致 {file_version}，跳过同步")
except FileNotFoundError:
    log_alert("WARNING", "筛选条件", "筛选条件.xlsx 不存在，跳过校验")
except Exception as e:
    log_alert("WARNING", "筛选条件", f"版本校验/同步失败: {str(e)[:80]}，继续推送")
# === 校验结束，开始推送 ===

repo_url = "https://github.com/lc132/lv.git"
repo_dir = "/tmp/lv_sync"
try:
    # 使用 GIT_ASKPASS 安全传递 Token（避免 Token 出现在进程列表中）
    import tempfile
    askpass_script = None
    try:
        fd, askpass_script = tempfile.mkstemp(prefix='git_askpass_', suffix='.sh')
        with os.fdopen(fd, 'w') as f:
            f.write('#!/bin/bash\necho "$GIT_TOKEN"\n')
        os.chmod(askpass_script, 0o700)
        git_env = os.environ.copy()
        git_env['GIT_ASKPASS'] = askpass_script
        git_env['GIT_TOKEN'] = token
        subprocess.run(
            ["git", "clone", "--depth", "1", "--branch", "main", repo_url, repo_dir],
            capture_output=True, text=True, timeout=30, check=True, env=git_env
        )
    finally:
        if askpass_script and os.path.exists(askpass_script):
            os.remove(askpass_script)
    # 推送筛选结果
    shutil.copy(md_path, os.path.join(repo_dir, f"短线标的_{prediction_date}.md"))
    # 若筛选条件表格已同步，一并推送
    if cond_synced and os.path.exists(cond_xlsx):
        shutil.copy(cond_xlsx, os.path.join(repo_dir, "A股短线选股筛选条件.xlsx"))
    subprocess.run(["git", "-C", repo_dir, "config", "user.email", "ashare-bot@github.com"], check=True)
    subprocess.run(["git", "-C", repo_dir, "config", "user.name", "ashare-screener"], check=True)
    subprocess.run(["git", "-C", repo_dir, "add", f"短线标的_{prediction_date}.md"], check=True)
    if cond_synced and os.path.exists(cond_xlsx):
        subprocess.run(["git", "-C", repo_dir, "add", "A股短线选股筛选条件.xlsx"], check=True)
    commit_msg = f"筛选结果 {prediction_date}"
    if cond_synced and xlsx_version and str(xlsx_version) != str(file_version):
        commit_msg += f" + 筛选条件同步至 {file_version}"
    subprocess.run(["git", "-C", repo_dir, "commit", "-m", commit_msg], check=True)
    result = subprocess.run(
        ["git", "-C", repo_dir, "push", "origin", "main"],
        capture_output=True, text=True, timeout=30
    )
    if result.returncode == 0:
        log_alert("INFO", "GitHub同步", f"✅ {prediction_date} 已推送")
    else:
        log_alert("WARNING", "GitHub同步", f"推送失败: {result.stderr[:100]}")
except Exception as e:
    log_alert("WARNING", "GitHub同步", f"失败: {str(e)[:100]}")
finally:
    if os.path.exists(repo_dir):
        shutil.rmtree(repo_dir, ignore_errors=True)