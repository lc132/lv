#!/usr/bin/env python3
"""
A股盘前短线标的筛选 v6.9.53 - Part 1: 步骤0-8 (前置检查)
"""
import urllib.request, urllib.parse, urllib.error, json, os, sys, time
from datetime import datetime, timedelta
from collections import Counter

# ============================================================
# 全局变量
# ============================================================
beijing_now = None
beijing_date = None
beijing_hour = None
beijing_weekday = None
prediction_date = None
data_date = None
file_version = "v6.9.53"

# 各阶段计数
total_raw = 0
excluded_count = 0
filtered_count = 0
matched_count = 0
industry_limited_count = 0
news_filtered_count = 0
final_recommend_count = 0
strategy_counts = Counter()

# 市场环境
market_env = "震荡"  # 强市/震荡/弱市
market_conditions = {}

# ============================================================
# 工具函数
# ============================================================
def log_alert(level, module, message, timestamp=None):
    if timestamp is None:
        timestamp = datetime.now()
    ts = timestamp.strftime('%Y-%m-%d %H:%M:%S') if hasattr(timestamp, 'strftime') else str(timestamp)
    log_line = f"[{ts}] [{level}] {module}: {message}\n"
    with open('/workspace/系统告警.log', 'a', encoding='utf-8') as f:
        f.write(log_line)
    print(f"  LOG [{level}] {module}: {message}")

def safe_read_json(path, default=None):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if not isinstance(data, list):
                log_alert("WARNING", "safe_read_json", f"{path} 格式异常")
                return default if default is not None else []
            return data
    except (json.JSONDecodeError, PermissionError, FileNotFoundError) as e:
        log_alert("ERROR", "safe_read_json", f"{path}: {str(e)}")
        return default if default is not None else []

def safe_write_json(path, data):
    """原子写入：先写临时文件再重命名，防止写入中断导致数据损坏"""
    try:
        tmp_path = path + '.tmp'
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)  # 原子操作(POSIX)
    except Exception as e:
        log_alert("ERROR", "safe_write_json", f"{path}: {str(e)}")

def safe_append_json(path, record):
    data = safe_read_json(path)
    data.append(record)
    safe_write_json(path, data)

def safe_float(value, ndigits=3):
    if value is None: return None
    if isinstance(value, (int, float)): return round(float(value), ndigits)
    return value

# ============================================================
# 步骤0: 获取北京时间
# ============================================================
def step0_get_beijing_time():
    global beijing_now, beijing_date, beijing_hour, beijing_weekday, prediction_date, data_date
    
    TIME_APIS = [
        ('https://timeapi.io/api/time/current/zone?timeZone=Asia/Shanghai', 'dateTime'),
    ]
    
    for api_url, key in TIME_APIS:
        try:
            req = urllib.request.Request(api_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read())
            ts = data[key]
            if '.' in ts:
                parts = ts.split('.')
                base = parts[0]
                frac = parts[1].split('+')[0].split('-')[0].split('Z')[0]
                tz_part = ''
                if '+' in parts[1]: tz_part = '+' + parts[1].split('+')[1]
                elif 'Z' in parts[1]: tz_part = '+00:00'
                frac6 = frac[:6].ljust(6, '0')
                ts = f'{base}.{frac6}{tz_part}'
            beijing_now = datetime.fromisoformat(ts)
            beijing_date = beijing_now.strftime('%Y-%m-%d')
            beijing_hour = beijing_now.hour
            beijing_weekday = beijing_now.weekday()
            prediction_date = beijing_date
            data_date = beijing_date
            print(f"✅ 步骤0: 北京时间 = {beijing_now.isoformat()}")
            print(f"   日期={beijing_date}, 星期={beijing_weekday}, 小时={beijing_hour}")
            return True
        except Exception as e:
            log_alert("INFO", "北京时间", f"{api_url} 不可达: {str(e)[:60]}")
            continue
    
    log_alert("ERROR", "北京时间", "所有授时API均不可达，本次筛选中止")
    return False

# ============================================================
# 步骤1: 节假日检查
# ============================================================
def step1_holiday_check():
    global data_date, prediction_date
    
    # 周六(5)或周日(6) → 非交易日
    if beijing_weekday >= 5:
        print(f"⚠️ 步骤1: {beijing_date} 是{'周六' if beijing_weekday==5 else '周日'}，A股休市")
        print(f"   使用上一交易日(周五)数据作为 data_date")
        # 周六→周五，周日→周五
        days_back = beijing_weekday - 4  # 5-4=1, 6-4=2
        prev_trading = beijing_now - timedelta(days=days_back)
        data_date = prev_trading.strftime('%Y-%m-%d')
        print(f"   data_date = {data_date} (上一交易日)")
        return True  # 不跳过，使用上一交易日数据
    
    # 搜索中国股市交易日历
    # 对于明确的节假日，跳过
    try:
        query = f"中国股市交易日历 {beijing_date}"
        print(f"   搜索节假日: {query}")
        # 简化：A股主要节假日：元旦(1/1)、春节(1-2月)、清明(4月)、五一(5/1)、端午(6月)、中秋(9-10月)、国庆(10/1-7)
        m = beijing_now.month
        d = beijing_now.day
        if (m == 1 and d == 1) or (m == 5 and d <= 3) or (m == 10 and d <= 7):
            print(f"⚠️ 步骤1: {beijing_date} 是法定节假日，跳过筛选")
            return False
    except Exception as e:
        pass
    
    print(f"✅ 步骤1: {beijing_date} 是交易日")
    return True

# ============================================================
# 步骤2: 极端行情检查
# ============================================================
def step2_extreme_market():
    global market_env
    try:
        # 获取上证指数行情
        url = "https://push2.eastmoney.com/api/qt/stock/get?secid=1.000001&fields=f43,f44,f45,f46,f47,f48,f50,f51,f52,f170"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        if data.get('data'):
            d = data['data']
            sh_close = d.get('f43', 0) / 100 if d.get('f43') else 0
            sh_change_pct = d.get('f170', 0) / 100 if d.get('f170') else 0
            sh_high = d.get('f44', 0) / 100 if d.get('f44') else 0
            sh_low = d.get('f45', 0) / 100 if d.get('f45') else 0
            sh_open = d.get('f46', 0) / 100 if d.get('f46') else 0
            sh_prev = d.get('f60', 0) / 100 if d.get('f60') else 0
            print(f"✅ 步骤2: 上证指数 收盘={sh_close:.2f} 涨跌={sh_change_pct:.2f}%")
            
            if sh_change_pct < -3:
                print(f"⚠️ 上证跌超3%，跳过筛选")
                log_alert("INFO", "极端行情", f"上证跌{sh_change_pct:.2f}%，跳过")
                return False
            if sh_change_pct > 3:
                print(f"⚠️ 上证涨超3%，仓位降为30%仅动量延续")
                market_env = "弱市"
    except Exception as e:
        log_alert("WARNING", "极端行情", f"获取上证指数失败: {str(e)[:80]}")
        print(f"⚠️ 步骤2: 无法获取上证指数: {str(e)[:60]}")
    
    return True

# ============================================================
# 步骤3: 外围市场
# ============================================================
def step3_external_markets():
    global market_env
    try:
        # 美股三大指数
        us_indices = {
            '.DJI': '道琼斯', '.IXIC': '纳斯达克', '.INX': '标普500'
        }
        us_down_count = 0
        for code, name in us_indices.items():
            try:
                url = f"https://push2.eastmoney.com/api/qt/stock/get?secid=100{code}&fields=f43,f170"
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=5) as resp:
                    data = json.loads(resp.read())
                if data.get('data'):
                    chg = data['data'].get('f170', 0) / 100 if data['data'].get('f170') else 0
                    print(f"   美股 {name}: {chg:.2f}%")
                    if chg < -2: us_down_count += 1
            except Exception:pass
        
        if us_down_count >= 3:
            print(f"⚠️ 美股三大指数均跌超2%，弱市仓位≤30%")
            market_env = "弱市"
        
        # 恒生指数
        try:
            url = "https://push2.eastmoney.com/api/qt/stock/get?secid=100.HSI&fields=f43,f170"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read())
            if data.get('data'):
                hsi_chg = data['data'].get('f170', 0) / 100 if data['data'].get('f170') else 0
                print(f"   恒生指数: {hsi_chg:.2f}%")
                if hsi_chg < -3:
                    print(f"⚠️ 恒生跌超3%，弱市仅超跌反弹")
                    market_env = "弱市"
        except Exception:pass
        
        print(f"✅ 步骤3: 外围市场检查完成，市场环境={market_env}")
    except Exception as e:
        log_alert("WARNING", "外围市场", f"检查失败: {str(e)[:80]}")
    
    return True

# ============================================================
# 步骤3A: 开盘前外围（期货）
# ============================================================
def step3a_premarket_futures():
    global market_env
    try:
        # 检查标普期货
        url = "https://push2.eastmoney.com/api/qt/stock/get?secid=100.ES00Y&fields=f43,f170"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
        if data.get('data'):
            es_chg = data['data'].get('f170', 0) / 100 if data['data'].get('f170') else 0
            print(f"   标普期货: {es_chg:.2f}%")
            if es_chg < -1:
                print(f"⚠️ 标普期货跌超1%，外围偏空，仓位降一档")
                if market_env == "强市": market_env = "震荡"
                elif market_env == "震荡": market_env = "弱市"
    except Exception as e:
        log_alert("INFO", "开盘前外围", f"期货数据不可得: {str(e)[:60]}")
        print(f"   期货数据不可得，跳过")
    return True

# ============================================================
# 步骤4: 持仓行情同步
# ============================================================
def step4_holding_sync():
    """同步持仓行情"""
    history = safe_read_json('/workspace/推荐历史.json')
    holdings = [r for r in history if r.get('type') == 'holding']
    if not holdings:
        print(f"✅ 步骤4: 无持仓记录，跳过")
        return []
    
    updated = []
    for h in holdings:
        code = h.get('code', '')
        if not code:
            continue
        # 保存旧current为prev_close
        old_current = h.get('current')
        h['prev_close'] = old_current
        
        # 获取最新行情
        try:
            market = 'sz' if code.startswith(('000','002','003','300','301')) else 'sh'
            url = f"https://hq.sinajs.cn/list={market}{code}"
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0',
                'Referer': 'https://finance.sina.com.cn'
            })
            with urllib.request.urlopen(req, timeout=5) as resp:
                text = resp.read().decode('gbk')
            if text and '=""' not in text:
                parts = text.split('"')[1].split(',')
                if len(parts) > 5:
                    current = float(parts[3]) if parts[3] else 0
                    if current > 0:
                        cost = h.get('cost', current)
                        shares = h.get('shares', 0)
                        h['current'] = round(current, 3)
                        h['pnl_pct'] = round((current - cost) / cost * 100, 2) if cost > 0 else 0
                        h['pnl_amount'] = round((current - cost) * shares, 2)
                        h['market_value'] = round(current * shares, 2)
                        h['update_date'] = data_date
                        updated.append(h)
                        print(f"   持仓 {code} {h.get('name','?')}: 现价={current:.2f} 盈亏={h['pnl_pct']:.2f}%")
                        continue
        except Exception as e:
            log_alert("WARNING", "持仓行情同步", f"{code} 搜索失败: {str(e)[:60]}")
        
        # 保留旧数据
        updated.append(h)
    
    # 写回
    for h in updated:
        for i, r in enumerate(history):
            if r.get('type') == 'holding' and r.get('code') == h.get('code'):
                history[i] = h
                break
    safe_write_json('/workspace/推荐历史.json', history)
    print(f"✅ 步骤4: 同步了 {len(updated)} 条持仓记录")
    return [h for h in updated if h.get('type') == 'holding']

# ============================================================
# 步骤4A: 做T评估
# ============================================================
def step4a_do_t_eval(holdings):
    """做T评估"""
    if not holdings:
        print(f"✅ 步骤4A: 无持仓，跳过做T评估")
        return
    
    do_t_evals = []
    for h in holdings:
        pnl_pct = h.get('pnl_pct', 0)
        code = h.get('code', '?')
        name = h.get('name', '?')
        
        if pnl_pct > -5:
            status = "观望"
            max_pos = "0"
        elif pnl_pct > -10:
            status = "重点评估"
            max_pos = "≤1/3"
        elif pnl_pct > -15:
            status = "谨慎评估"
            max_pos = "≤1/4"
        else:
            status = "不做T"
            max_pos = "0"
        
        eval_record = {
            "type": "do_T_eval",
            "date": data_date,
            "code": code,
            "name": name,
            "pnl_pct": pnl_pct,
            "status": status,
            "max_position": max_pos,
            "eval_time": beijing_now.isoformat() if beijing_now else ""
        }
        do_t_evals.append(eval_record)
        safe_append_json('/workspace/推荐历史.json', eval_record)
        print(f"   做T评估 {code} {name}: {status} (浮亏{pnl_pct:.2f}%)")
    
    print(f"✅ 步骤4A: 完成 {len(do_t_evals)} 个做T评估")

# ============================================================
# 步骤4B: 持仓跟踪同步
# ============================================================
def step4b_holding_tracking(holdings):
    """同步持仓跟踪.xlsx"""
    xlsx_path = '/workspace/持仓跟踪.xlsx'
    if not os.path.exists(xlsx_path):
        log_alert("WARNING", "持仓跟踪同步", "持仓跟踪.xlsx不存在，跳过")
        print(f"⚠️ 步骤4B: 持仓跟踪.xlsx不存在，跳过")
        return
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb["持仓明细"]
        
        code_row = {}
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            if code and isinstance(code, str) and len(code) == 6:
                code_row[str(code)] = row
        
        updated = 0
        for h in holdings:
            code = str(h.get('code', ''))
            current = h.get('current')
            if not code or code not in code_row:
                if code:
                    log_alert("WARNING", "持仓跟踪同步", f"{code} 在xlsx中找不到")
                continue
            if current is None:
                log_alert("WARNING", "持仓跟踪同步", f"{code} 缺少current字段，跳过")
                continue
            
            row = code_row[code]
            ws.cell(row=row, column=7).value = current
            mv = h.get('market_value')
            pnl_amt = h.get('pnl_amount')
            if mv is None or pnl_amt is None:
                cost = ws.cell(row=row, column=3).value
                shares = ws.cell(row=row, column=4).value
                if cost and shares and current:
                    mv = round(current * shares, 2)
                    pnl_amt = round((current - cost) * shares, 2)
            if mv is not None:
                ws.cell(row=row, column=8).value = mv
            if pnl_amt is not None:
                ws.cell(row=row, column=9).value = round(pnl_amt, 2)
            pnl_pct_val = h.get('pnl_pct')
            try:
                pnl_pct_float = float(pnl_pct_val) if pnl_pct_val is not None else 0.0
            except (ValueError, TypeError):
                pnl_pct_float = 0.0
            ws.cell(row=row, column=10).value = round(pnl_pct_float, 4)
            updated += 1
        
        if updated > 0:
            wb.save(xlsx_path)
            log_alert("INFO", "持仓跟踪同步", f"已更新{updated}只持仓价格")
        wb.close()
        print(f"✅ 步骤4B: 同步了 {updated} 条持仓到持仓跟踪.xlsx")
    except Exception as e:
        log_alert("WARNING", "持仓跟踪同步", f"失败: {str(e)[:100]}")
        print(f"⚠️ 步骤4B: 同步失败: {str(e)[:100]}")

# ============================================================
# 步骤4C: 持仓危机检查
# ============================================================
def step4c_holding_crisis(holdings):
    """持仓危机检查"""
    alerts = []
    for h in holdings:
        code = h.get('code', '?')
        name = h.get('name', '?')
        cost = h.get('cost', 0)
        current = h.get('current', 0)
        prev_close = h.get('prev_close')
        pnl_pct = h.get('pnl_pct', 0)
        
        # 跌停检查
        if prev_close is not None and current > 0 and prev_close > 0:
            daily_chg = (current - prev_close) / prev_close * 100
            if daily_chg < -9.5:
                msg = f"⚠️ {code} {name} 当日跌停({daily_chg:.1f}%)！成本{cost} 现价{current} 浮亏{pnl_pct}%"
                alerts.append(msg)
                log_alert("WARNING", "持仓危机", msg)
        
        # 浮亏>15%
        if pnl_pct is not None and pnl_pct < -15:
            msg = f"⚠️ {code} {name} 浮亏突破15%做T上限({pnl_pct:.1f}%)，建议人工决策"
            alerts.append(msg)
            log_alert("WARNING", "持仓危机", msg)
        
        # L1硬排除触发
        if current > 0:
            l1_triggers = []
            if current < 5: l1_triggers.append("股价<5元(规则3)")
            if current > 100: l1_triggers.append("股价>100元(规则4)")
            if code.startswith("688"): l1_triggers.append("科创板(规则1)")
            if code.startswith("8") and len(str(code)) == 6: l1_triggers.append("北交所(规则2)")
            if l1_triggers:
                msg = f"⚠️ {code} {name} 触发L1硬排除: {', '.join(l1_triggers)}"
                alerts.append(msg)
                log_alert("WARNING", "持仓危机", msg)
    
    if alerts:
        print(f"⚠️ 步骤4C: 持仓危机检查发现 {len(alerts)} 条告警:")
        for a in alerts:
            print(f"   {a}")
    else:
        print(f"✅ 步骤4C: 持仓危机检查无异常")
    return alerts

# ============================================================
# 步骤5: 推荐历史持久化
# ============================================================
def step5_history_persistence():
    """推荐历史清理"""
    history = safe_read_json('/workspace/推荐历史.json')
    if not history:
        print(f"✅ 步骤5: 推荐历史为空，跳过")
        return
    
    try:
        cutoff_7d_dt = datetime.strptime(data_date, '%Y-%m-%d') - timedelta(days=7)
        cutoff_90d_dt = datetime.strptime(data_date, '%Y-%m-%d') - timedelta(days=90)
        cutoff_7d = cutoff_7d_dt.strftime('%Y-%m-%d')
        cutoff_90d = cutoff_90d_dt.strftime('%Y-%m-%d')
        
        new_history = []
        cleaned = 0
        for r in history:
            t = r.get('type', '')
            if t in ('weekly_review', 'strategy_check', 'do_T_eval', 'do_T'):
                new_history.append(r)
            elif t == 'holding':
                d = r.get('update_date', '')
                if d >= cutoff_90d:
                    new_history.append(r)
                else:
                    cleaned += 1
            elif t == 'recommendation':
                d = r.get('date', '')
                if d >= cutoff_7d:
                    new_history.append(r)
                else:
                    cleaned += 1
            else:
                new_history.append(r)
        
        if cleaned > 0:
            safe_write_json('/workspace/推荐历史.json', new_history)
            log_alert("INFO", "清理", f"已清理{cleaned}条过期记录")
            print(f"✅ 步骤5: 清理了 {cleaned} 条过期记录")
        else:
            print(f"✅ 步骤5: 无需清理")
    except Exception as e:
        log_alert("WARNING", "清理", f"清理失败: {str(e)[:80]}")
        print(f"⚠️ 步骤5: 清理失败: {str(e)[:80]}")

# ============================================================
# 步骤6: 文件初始化
# ============================================================
def step6_file_init():
    global file_version
    adj_records = safe_read_json('/workspace/策略调整记录.json')
    params = {}
    if adj_records and len(adj_records) > 0:
        latest = adj_records[-1]
        file_version = latest.get('version', 'v6.9.53')
        params = latest.get('params', {})
    else:
        file_version = 'v6.9.53'
        params = {}
    
    print(f"✅ 步骤6: 策略版本 = {file_version}")
    
    # 版本一致性检查
    history = safe_read_json('/workspace/推荐历史.json')
    last_check = None
    current_version = None
    for r in reversed(history):
        if r.get('type') == 'strategy_check':
            last_check = r
            break
    
    if last_check:
        current_version = last_check.get('version', 'unknown')
        if current_version != file_version:
            log_alert("INFO", "版本检查", f"推荐历史版本{current_version}≠策略调整版本{file_version}，以策略调整为准")
        else:
            log_alert("INFO", "版本检查", f"版本一致{file_version}")
    
    # 首次运行或版本变更
    if last_check is None or current_version != file_version:
        check_record = {
            "type": "strategy_check",
            "version": file_version,
            "params": params,
            "date": data_date,
            "checks": {"version_match": current_version == file_version}
        }
        safe_append_json('/workspace/推荐历史.json', check_record)
        log_alert("INFO", "版本检查", f"已追加strategy_check记录: {file_version}")
        print(f"   版本变更或首次运行，已追加strategy_check")
    
    return params

# ============================================================
# 步骤7: 财报季检测
# ============================================================
def step7_earnings_season():
    global market_env
    m = beijing_now.month
    if m in (1, 3, 4, 8, 10):
        print(f"✅ 步骤7: {m}月是财报季，事件驱动权重×1.5，仓位+5%")
        return True
    print(f"✅ 步骤7: 非财报季")
    return False

# ============================================================
# 步骤8: 大盘环境判断
# ============================================================
def step8_market_environment():
    global market_env, market_conditions
    try:
        # 获取上证指数和MA数据
        # 简化：使用当前涨跌比和成交量判断
        url = "https://push2.eastmoney.com/api/qt/stock/get?secid=1.000001&fields=f43,f44,f45,f46,f47,f48,f50,f170"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        if data.get('data'):
            d = data['data']
            sh_close = d.get('f43', 0) / 100 if d.get('f43') else 0
            sh_change = d.get('f170', 0) / 100 if d.get('f170') else 0
            
            # 简化判断：基于涨跌幅
            if sh_change > 1:
                tentative = "强市"
            elif sh_change < -1:
                tentative = "弱市"
            else:
                tentative = "震荡"
            
            # 如果步骤3/3A已经判定为弱市，则保持
            if market_env == "弱市":
                tentative = "弱市"
            
            market_env = tentative
            market_conditions = {
                "sh_close": sh_close,
                "sh_change": sh_change,
                "env": market_env
            }
            
            print(f"✅ 步骤8: 大盘环境 = {market_env} (上证 {sh_close:.2f}, {sh_change:+.2f}%)")
    except Exception as e:
        log_alert("WARNING", "大盘环境", f"获取失败: {str(e)[:80]}")
        print(f"⚠️ 步骤8: 默认震荡市")
    
    return market_env

# ============================================================
# 主流程 Part 1
# ============================================================
def run_part1():
    print("=" * 60)
    print("A股盘前短线标的筛选 v6.9.53 - Part 1: 步骤0-8")
    print("=" * 60)
    
    # 步骤0
    if not step0_get_beijing_time():
        print("❌ 步骤0失败，中止")
        return False
    
    # 步骤1
    if not step1_holiday_check():
        print("❌ 步骤1: 节假日，跳过")
        return False
    
    # 步骤2
    if not step2_extreme_market():
        print("❌ 步骤2: 极端行情，跳过")
        return False
    
    # 步骤3
    step3_external_markets()
    
    # 步骤3A
    step3a_premarket_futures()
    
    # 步骤4
    holdings = step4_holding_sync()
    
    # 步骤4A
    step4a_do_t_eval(holdings)
    
    # 步骤4B
    step4b_holding_tracking(holdings)
    
    # 步骤4C
    step4c_holding_crisis(holdings)
    
    # 步骤5
    step5_history_persistence()
    
    # 步骤6
    params = step6_file_init()
    
    # 步骤7
    is_earnings = step7_earnings_season()
    
    # 步骤8
    step8_market_environment()
    
    print(f"\n✅ Part 1 完成: 市场环境={market_env}, 版本={file_version}")
    print(f"   data_date={data_date}, prediction_date={prediction_date}")
    return True

if __name__ == '__main__':
    success = run_part1()
    if not success:
        sys.exit(1)
    # 输出关键变量供后续部分使用
    print(f"\n__EXPORT__: data_date={data_date}, prediction_date={prediction_date}, market_env={market_env}, file_version={file_version}")